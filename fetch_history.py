"""
fetch_history.py — Pull complete bet history for all sharp wallets and export to Excel.

Fetches every resolved and open position for each tracked wallet, classifies
wins/losses, groups by sport, and writes a color-coded workbook.

Usage:
    py fetch_history.py                         # all wallets, all sports
    py fetch_history.py --sports Soccer MLB NBA  # filter to specific sports
    py fetch_history.py --wallets Jargs swisstony
    py fetch_history.py --limit 1000            # positions per wallet (default 500)
    py fetch_history.py --out my_report.xlsx    # custom output filename
"""

import argparse
import io
import json
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

POSITIONS_API    = "https://data-api.polymarket.com/positions"
SHARP_FILE       = os.getenv("SHARP_WALLETS_FILE", "sharp_wallets_selected.json")
DEFAULT_PER_PAGE = 500
MAX_PAGES        = 10        # safety cap — 10 × 500 = 5 000 positions per wallet max

CATEGORY_RULES = [
    ("wnba",              "WNBA"),
    ("nba",               "NBA"),
    ("nfl",               "NFL"),
    ("mlb",               "MLB"),
    ("nhl",               "NHL"),
    ("fifwc",             "FIFA WC"),
    ("world-cup",         "FIFA WC"),
    ("world cup",         "FIFA WC"),
    ("epl",               "Soccer"),
    ("ucl",               "Soccer"),
    ("laliga",            "Soccer"),
    ("premier-league",    "Soccer"),
    ("premier league",    "Soccer"),
    ("champions-league",  "Soccer"),
    ("champions league",  "Soccer"),
    ("mls",               "Soccer"),
    ("serie-a",           "Soccer"),
    ("bundesliga",        "Soccer"),
    ("ligue-1",           "Soccer"),
    ("ufc",               "UFC/MMA"),
    ("mma",               "UFC/MMA"),
    ("bellator",          "UFC/MMA"),
    ("tennis",            "Tennis"),
    ("wimbledon",         "Tennis"),
    ("atp",               "Tennis"),
    ("wta",               "Tennis"),
    ("us-open",           "Tennis"),
    ("french-open",       "Tennis"),
    ("australian-open",   "Tennis"),
    ("formula-1",         "F1"),
    ("formula1",          "F1"),
    ("grand-prix",        "F1"),
    ("pga",               "Golf"),
    (" golf",             "Golf"),
    ("masters",           "Golf"),
    ("ryder",             "Golf"),
    ("ncaab",             "NCAAB"),
    ("march-madness",     "NCAAB"),
    ("ncaaf",             "NCAAF"),
    ("college-football",  "NCAAF"),
    ("valorant",          "Esports"),
    ("csgo",              "Esports"),
    ("cs2",               "Esports"),
    ("league-of-legends", "Esports"),
    ("dota",              "Esports"),
    ("boxing",            "Boxing"),
    ("cricket",           "Cricket"),
]


def classify(title: str, slug: str) -> str:
    text = f"{slug} {title}".lower()
    for kw, cat in CATEGORY_RULES:
        if kw in text:
            return cat
    return "Other"


def as_float(v, default=0.0):
    try:
        return float(v or default)
    except (TypeError, ValueError):
        return default


def american_odds(cents: float) -> str:
    p = as_float(cents) / 100
    if p <= 0 or p >= 1:
        return "n/a"
    if p >= 0.5:
        return str(round(-100 * p / (1 - p)))
    return f"+{round(100 * (1 - p) / p)}"


def position_status(pos: dict) -> str:
    if pos.get("redeemable"):
        return "WIN"
    cur = as_float(pos.get("curPrice"))
    if cur <= 0:
        return "LOSS"
    return "OPEN"


def fetch_wallet_positions(addr: str, label: str, limit: int) -> list[dict]:
    """Fetch all positions for one wallet with offset pagination."""
    results = []
    offset  = 0
    session = requests.Session()
    session.headers["User-Agent"] = "polytrack-history/1.0"

    for _ in range(MAX_PAGES):
        try:
            r = session.get(
                POSITIONS_API,
                params={"user": addr, "limit": limit, "offset": offset},
                timeout=20,
            )
            if r.status_code != 200:
                print(f"  [{label}] HTTP {r.status_code} at offset {offset}")
                break
            page = r.json()
            if not page:
                break
            results.extend(page)
            if len(page) < limit:
                break
            offset += limit
            time.sleep(0.1)
        except Exception as exc:
            print(f"  [{label}] fetch error: {exc}")
            break

    return results


def build_row(addr: str, label: str, cat: str, pos: dict) -> dict:
    initial  = as_float(pos.get("initialValue"))
    cur_val  = as_float(pos.get("currentValue"))
    cash_pnl = as_float(pos.get("cashPnl"))
    avg_p    = round(as_float(pos.get("avgPrice")) * 100, 2)
    cur_p    = round(as_float(pos.get("curPrice"))  * 100, 2)
    status   = position_status(pos)

    if status == "WIN":
        pl = cash_pnl if cash_pnl != 0 else (cur_val - initial)
    elif status == "LOSS":
        pl = cash_pnl if cash_pnl != 0 else -initial
    else:
        pl = cash_pnl

    roi = pl / initial * 100 if initial > 0 else 0.0
    slug = pos.get("eventSlug") or ""

    return {
        "wallet":   label,
        "address":  addr,
        "sport":    cat,
        "market":   pos.get("title", "?"),
        "outcome":  pos.get("outcome", "?"),
        "status":   status,
        "avg_p":    avg_p,
        "cur_p":    cur_p,
        "entry_odds": american_odds(avg_p),
        "cur_odds":   american_odds(cur_p),
        "invested": round(initial, 2),
        "cur_value": round(cur_val, 2),
        "pl":       round(pl, 2),
        "roi":      round(roi, 2),
        "shares":   round(as_float(pos.get("size") or pos.get("shares")), 4),
        "end_date": (pos.get("endDate") or "")[:10],
        "event_slug": slug,
        "source":   f"https://polymarket.com/event/{slug}" if slug else "",
    }


# ── Excel helpers ─────────────────────────────────────────────────────────────

def fill(hex6):
    return PatternFill("solid", fgColor="FF" + hex6.lstrip("#").upper())

def fnt(hex6, bold=False, size=10):
    return Font(color="FF" + hex6.lstrip("#").upper(), bold=bold, size=size)

HDR_FILL = fill("0D1F16")
HDR_FONT = fnt("34C759", bold=True, size=10)
WIN_FILL = fill("0A1F12")
LOS_FILL = fill("1F0D0D")
OPN_FILL = fill("111814")
YES_FONT = fnt("34C759", bold=True)
NO_FONT  = fnt("F0606E", bold=True)
POS_FONT = fnt("34C759", bold=True)
NEG_FONT = fnt("F0606E", bold=True)
GLD_FONT = fnt("E6A817", bold=True)
WHT_FONT = fnt("D8EBE0")
DIM_FONT = fnt("5A7A68")
GRN_FONT = fnt("34C759")
RED_FONT = fnt("F0606E")
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center")


def make_sheet(wb: Workbook, title: str, rows: list[dict]):
    ws = wb.create_sheet(title[:31])

    headers = [
        ("Wallet",      14), ("Sport",       11), ("Market",      46),
        ("Side",         6), ("Status",       7), ("Entry Odds",  11),
        ("Cur/Fin Odds", 12), ("Invested $",  11), ("Shares",       9),
        ("P&L $",       11), ("ROI %",         8), ("End Date",    11),
        ("Link",        10),
    ]

    for col, (h, w) in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    for ri, r in enumerate(rows, 2):
        st  = r["status"]
        rf  = WIN_FILL if st == "WIN" else LOS_FILL if st == "LOSS" else OPN_FILL
        sf  = (fnt("34C759", bold=True) if st == "WIN"
               else fnt("F0606E", bold=True) if st == "LOSS"
               else GLD_FONT)
        pf  = POS_FONT if r["pl"] > 0 else NEG_FONT if r["pl"] < 0 else WHT_FONT
        rf2 = POS_FONT if r["roi"] > 0 else NEG_FONT if r["roi"] < 0 else WHT_FONT
        of  = YES_FONT if r["outcome"] == "Yes" else NO_FONT

        for ci, (v, fn, al, fmt) in enumerate([
            (r["wallet"],        WHT_FONT, LEFT,   None),
            (r["sport"],         DIM_FONT, CENTER, None),
            (r["market"],        WHT_FONT, LEFT,   None),
            (r["outcome"],       of,       CENTER, None),
            (st,                 sf,       CENTER, None),
            (r["entry_odds"],    DIM_FONT, CENTER, None),
            (r["cur_odds"],      WHT_FONT, CENTER, None),
            (r["invested"],      DIM_FONT, CENTER, "#,##0.00"),
            (r["shares"],        DIM_FONT, CENTER, "0.0000"),
            (r["pl"],            pf,       CENTER, "#,##0.00"),
            (r["roi"],           rf2,      CENTER, "0.00"),
            (r["end_date"],      DIM_FONT, CENTER, None),
            (r["source"],        DIM_FONT, LEFT,   None),
        ], 1):
            c2 = ws.cell(ri, ci, v)
            c2.fill = rf
            c2.font = fn
            c2.alignment = al
            if fmt: c2.number_format = fmt
        ws.row_dimensions[ri].height = 16

    return ws


def make_summary_sheet(wb: Workbook, all_rows: list[dict], wallets_meta: dict):
    ws = wb.active
    ws.title = "Summary"

    # ── per-wallet totals ──────────────────────────────────────────────────────
    ws.cell(1, 1, "WALLET SUMMARY").font  = fnt("34C759", bold=True, size=12)
    ws.cell(1, 1).fill = fill("0D1F16")

    w_hdrs = [
        ("Wallet", 16), ("Sport", 14), ("Bets", 7), ("Wins", 6), ("Losses", 7),
        ("Open", 6), ("Win %", 8), ("Invested $", 12), ("P&L $", 12), ("ROI %", 8),
        ("Sharpness", 11),
    ]
    for col, (h, w) in enumerate(w_hdrs, 1):
        c = ws.cell(3, col, h)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

    # group by (wallet_label, sport)
    by_ws: dict = {}
    for r in all_rows:
        key = (r["wallet"], r["sport"])
        g   = by_ws.setdefault(key, {"wins": 0, "losses": 0, "open": 0, "inv": 0.0, "pl": 0.0})
        if r["status"] == "WIN":   g["wins"]   += 1
        elif r["status"] == "LOSS": g["losses"] += 1
        else:                       g["open"]   += 1
        g["inv"] += r["invested"]
        g["pl"]  += r["pl"]

    ri = 4
    for (wlabel, sport), g in sorted(by_ws.items(), key=lambda x: -x[1]["pl"]):
        total = g["wins"] + g["losses"]
        wr    = g["wins"] / total * 100 if total > 0 else 0.0
        roi   = g["pl"] / g["inv"] * 100 if g["inv"] > 0 else 0.0
        sharp = wallets_meta.get((wlabel, sport), {}).get("sharpness_score", 0)
        pf    = POS_FONT if g["pl"]  > 0 else NEG_FONT
        rf    = POS_FONT if roi > 0 else NEG_FONT if roi < 0 else WHT_FONT
        wrf   = (fnt("34C759", bold=True) if wr >= 55
                 else fnt("F0606E", bold=True) if wr < 50 and total >= 10
                 else WHT_FONT)
        for ci, (v, fn, fmt) in enumerate([
            (wlabel,                WHT_FONT, None),
            (sport,                 DIM_FONT, None),
            (g["wins"]+g["losses"]+g["open"], DIM_FONT, None),
            (g["wins"],             GRN_FONT, None),
            (g["losses"],           RED_FONT, None),
            (g["open"],             GLD_FONT, None),
            (round(wr, 1),          wrf,      "0.0"),
            (round(g["inv"], 2),    DIM_FONT, "#,##0.00"),
            (round(g["pl"], 2),     pf,       "#,##0.00"),
            (round(roi, 2),         rf,       "0.00"),
            (round(as_float(sharp), 1), GLD_FONT if as_float(sharp) >= 70 else WHT_FONT, "0.0"),
        ], 1):
            c = ws.cell(ri, ci, v)
            c.fill = OPN_FILL
            c.font = fn
            c.alignment = CENTER if ci > 1 else LEFT
            if fmt: c.number_format = fmt
        ws.row_dimensions[ri].height = 16
        ri += 1

    return ws


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Export full bet history for sharp wallets")
    parser.add_argument("--sports",  nargs="+", help="Limit to specific sport(s)")
    parser.add_argument("--wallets", nargs="+", help="Limit to specific wallet label(s)")
    parser.add_argument("--limit",   type=int, default=DEFAULT_PER_PAGE, help="Positions per page")
    parser.add_argument("--out",     default="", help="Output filename (default: history_YYYYMMDD_HHMM.xlsx)")
    parser.add_argument("--workers", type=int, default=8, help="Parallel wallet fetch threads")
    args = parser.parse_args()

    if not os.path.exists(SHARP_FILE):
        print(f"ERROR: {SHARP_FILE} not found. Run py ingest_sharp_wallets.py --no-near-term-filter first.")
        sys.exit(1)

    with open(SHARP_FILE, "r", encoding="utf-8") as f:
        payload = json.load(f)

    wallet_records = payload.get("wallets", [])
    wallets_meta: dict = {}  # (label, sport) → record
    wallets: list[tuple[str, str, str]] = []  # (addr, label, sport)

    for rec in wallet_records:
        addr  = (rec.get("wallet_address") or "").lower()
        label = rec.get("wallet_label") or addr[:10]
        sport = rec.get("category", "Other")

        if args.wallets and label not in args.wallets:
            continue
        if args.sports and sport not in args.sports:
            continue
        if not addr:
            continue

        wallets.append((addr, label, sport))
        wallets_meta[(label, sport)] = rec

    # deduplicate addresses — one fetch per address covers all sports
    addr_labels: dict[str, str] = {}
    for addr, label, _ in wallets:
        addr_labels[addr] = label

    print(f"Fetching history for {len(addr_labels)} unique wallets "
          f"({len(wallets)} wallet×sport combinations)…")

    raw_positions: dict[str, list] = {}

    def fetch_one(addr_label):
        addr, label = addr_label
        print(f"  → {label} ({addr[:10]}…)")
        return addr, fetch_wallet_positions(addr, label, args.limit)

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futures = {ex.submit(fetch_one, (addr, label)): addr
                   for addr, label in addr_labels.items()}
        for fut in as_completed(futures):
            addr, positions = fut.result()
            raw_positions[addr] = positions
            print(f"     {addr_labels[addr]}: {len(positions)} positions fetched")

    # ── build rows ─────────────────────────────────────────────────────────────
    # Determine which (addr, sport) combos we care about
    wanted_sports: dict[str, set[str]] = {}
    for addr, label, sport in wallets:
        wanted_sports.setdefault(addr, set()).add(sport)

    all_rows: list[dict] = []
    for addr, label, sport in wallets:
        for pos in raw_positions.get(addr, []):
            title = pos.get("title", "")
            slug  = pos.get("eventSlug", "") or ""
            cat   = classify(title, slug)
            # only include rows that match this wallet's tracked sport
            if cat != sport:
                continue
            all_rows.append(build_row(addr, label, sport, pos))

    if not all_rows:
        print("No positions found. Check wallet addresses and ensure the API is reachable.")
        sys.exit(0)

    total_bets = len(all_rows)
    wins  = sum(1 for r in all_rows if r["status"] == "WIN")
    losses= sum(1 for r in all_rows if r["status"] == "LOSS")
    open_ = sum(1 for r in all_rows if r["status"] == "OPEN")
    total_pl = sum(r["pl"] for r in all_rows)
    total_inv= sum(r["invested"] for r in all_rows)
    print(f"\n{total_bets} bets  |  {wins}W {losses}L {open_}Open  |  "
          f"${total_pl:,.0f} P&L on ${total_inv:,.0f} invested "
          f"({total_pl/total_inv*100:.1f}% ROI)" if total_inv > 0 else "")

    # ── build workbook ─────────────────────────────────────────────────────────
    wb = Workbook()

    make_summary_sheet(wb, all_rows, wallets_meta)

    # one sheet per sport, sorted by # bets desc
    by_cat: dict[str, list] = {}
    for r in all_rows:
        by_cat.setdefault(r["sport"], []).append(r)

    for sport, rows in sorted(by_cat.items(), key=lambda x: -len(x[1])):
        rows_sorted = sorted(rows, key=lambda r: r["end_date"], reverse=True)
        make_sheet(wb, sport, rows_sorted)

    fname = args.out or f"history_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    print(f"\nSaved → {fname}")
    print(f"Sheets: Summary + {', '.join(sorted(by_cat))}")


if __name__ == "__main__":
    main()
