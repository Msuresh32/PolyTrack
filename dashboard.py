"""
Polymarket position dashboard — Flask web app.

Fetches all open positions for tracked wallets from data-api.polymarket.com,
enriches them with conviction scores derived from the fills history in seen.db
(written by monitor.py), and serves a filterable table at http://localhost:5000.

Run alongside monitor.py (separate terminal):
    pip install flask requests
    python dashboard.py

The positions endpoint is:
    data-api.polymarket.com/positions?user={address}&limit=200

Key fields returned (confirmed live 2026-06-26):
    avgPrice, curPrice, initialValue (USDC spent), currentValue, cashPnl,
    percentPnl, title, outcome (Yes/No), endDate, eventSlug, conditionId
"""

import os
import json
import sqlite3
import statistics
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
from typing import Optional

import requests
from flask import Flask, jsonify, render_template_string
from dotenv import load_dotenv

load_dotenv()

# ── Config ─────────────────────────────────────────────────────────────────────
DB              = "seen.db"
POSITIONS_API   = "https://data-api.polymarket.com/positions"
ANALYTICS_API   = os.getenv(
    "POLYMARKET_ANALYTICS_API",
    "https://legacy.polymarketanalytics.com/api/traders-tag-performance",
)
ANALYTICS_STATS = os.getenv("POLYMARKET_ANALYTICS_STATS", "true").lower() == "true"
REFRESH_SECONDS = int(os.getenv("DASHBOARD_REFRESH", "60"))
PORT            = int(os.getenv("DASHBOARD_PORT", "5000"))
SHARP_WALLETS_FILE = os.getenv("SHARP_WALLETS_FILE", "sharp_wallets_selected.json")

WALLETS: dict = {}
for _pair in os.getenv("WALLETS", "").split(","):
    _pair = _pair.strip()
    if "=" in _pair:
        _label, _addr = _pair.split("=", 1)
        WALLETS[_addr.strip().lower()] = _label.strip()
MANUAL_WALLETS = set(WALLETS)
SHARP_WALLET_POOLS: dict = {}
SHARP_WALLET_META: dict = {}

MY_BASE_UNIT   = float(os.getenv("MY_BASE_UNIT", "50"))
RESOLVE_WINDOW = int(os.getenv("RESOLVE_WINDOW_DAYS", "2"))  # only show markets ending within N days
MIN_AMERICAN_ODDS = int(os.getenv("MIN_AMERICAN_ODDS", "-250"))
MAX_AMERICAN_ODDS = int(os.getenv("MAX_AMERICAN_ODDS", "250"))

# ── Category classifier (mirrors monitor.py) ───────────────────────────────────
CATEGORY_RULES = [
    ("wnba",              "WNBA"),
    ("nba",               "NBA"),
    ("nfl",               "NFL"),
    ("mlb",               "MLB"),
    ("nhl",               "NHL"),
    ("fifwc",             "FIFA WC"),
    ("world-cup",         "FIFA WC"),
    ("world cup",         "FIFA WC"),
    ("epl",               "Soccer"),
    ("ucl",               "Soccer"),
    ("laliga",            "Soccer"),
    ("premier-league",    "Soccer"),
    ("premier league",    "Soccer"),
    ("champions-league",  "Soccer"),
    ("champions league",  "Soccer"),
    ("mls",               "Soccer"),
    ("serie-a",           "Soccer"),
    ("bundesliga",        "Soccer"),
    ("ligue-1",           "Soccer"),
    ("ufc",               "UFC/MMA"),
    ("mma",               "UFC/MMA"),
    ("bellator",          "UFC/MMA"),
    ("tennis",            "Tennis"),
    ("wimbledon",         "Tennis"),
    ("atp",               "Tennis"),
    ("wta",               "Tennis"),
    ("us-open",           "Tennis"),
    ("french-open",       "Tennis"),
    ("australian-open",   "Tennis"),
    ("formula-1",         "F1"),
    ("formula1",          "F1"),
    ("grand-prix",        "F1"),
    ("pga",               "Golf"),
    (" golf",             "Golf"),
    ("masters",           "Golf"),
    ("ryder",             "Golf"),
    ("ncaab",             "NCAAB"),
    ("march-madness",     "NCAAB"),
    ("ncaaf",             "NCAAF"),
    ("college-football",  "NCAAF"),
    ("valorant",          "Esports"),
    ("csgo",              "Esports"),
    ("cs2",               "Esports"),
    ("league-of-legends", "Esports"),
    ("dota",              "Esports"),
    ("boxing",            "Boxing"),
]

CATEGORY_ANALYTICS_TAGS = {
    "FIFA WC": "FIFA World Cup",
    "UFC/MMA": "UFC",
    "F1": "Formula 1",
}


def load_selected_sharp_wallets(path: str = SHARP_WALLETS_FILE) -> None:
    if not path or not os.path.exists(path):
        return

    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except Exception as exc:
        print(f"[dashboard] sharp wallet load error: {exc}")
        return

    for record in payload.get("wallets", []):
        addr = (record.get("wallet_address") or "").lower()
        category = {
            "FIFA World Cup": "FIFA WC",
        }.get(record.get("category"), record.get("category"))
        if not addr or not category:
            continue
        label = record.get("wallet_label") or addr[:10]
        WALLETS.setdefault(addr, label)
        SHARP_WALLET_POOLS.setdefault(category, set()).add(addr)
        SHARP_WALLET_META[(addr, category)] = record


load_selected_sharp_wallets()


def selected_wallet_meta(addr: str, category: str) -> dict:
    return SHARP_WALLET_META.get(((addr or "").lower(), category), {})


def classify_category(title: str, event_slug: str) -> str:
    text = f"{event_slug} {title}".lower()
    for kw, cat in CATEGORY_RULES:
        if kw in text:
            return cat
    return "Other"


# ── Stats helpers (mirrors monitor.py — reads from same seen.db) ───────────────
def wallet_median(con: sqlite3.Connection, wallet: str) -> float:
    rows = con.execute(
        "SELECT usd_size FROM fills WHERE wallet=?", (wallet,)
    ).fetchall()
    if not rows:
        return 0.0
    return statistics.median(r[0] for r in rows)


def conviction_score(usd_size: float, median: float) -> int:
    """
    Mirror of monitor.py conviction_score:
        score = clamp(round((usd_size / median) / 3.0 * 100), 0, 100)
    Returns 50 when no fill history exists.
    """
    if median <= 0:
        return 50
    return min(100, max(0, round((usd_size / median) / 3.0 * 100)))


# ── Position fetching ──────────────────────────────────────────────────────────
def as_float(value, default: float = 0.0) -> float:
    try:
        return float(value or default)
    except (TypeError, ValueError):
        return default


def clamp(value: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, value))


def percentile_score(value: float, samples: list) -> int:
    values = sorted(as_float(v) for v in samples if as_float(v) > 0)
    if not values or value <= 0:
        return 50

    less = sum(1 for v in values if v < value)
    equal = sum(1 for v in values if v == value)
    return round(clamp((less + 0.5 * equal) / len(values) * 100))


def american_odds_from_cents(cents: float) -> str:
    p = as_float(cents) / 100
    if p <= 0 or p >= 1:
        return "n/a"
    if p >= 0.5:
        return str(round(-100 * p / (1 - p)))
    return f"+{round(100 * (1 - p) / p)}"


def american_odds_value_from_cents(cents: float) -> Optional[int]:
    odds = american_odds_from_cents(cents)
    if odds == "n/a":
        return None
    return int(odds.replace("+", ""))


def cents_from_american_odds(odds: int) -> float:
    if odds < 0:
        return abs(odds) / (abs(odds) + 100) * 100
    return 100 / (odds + 100) * 100


def odds_in_range(cents: float) -> bool:
    low_cents = cents_from_american_odds(MAX_AMERICAN_ODDS)
    high_cents = cents_from_american_odds(MIN_AMERICAN_ODDS)
    return low_cents <= cents <= high_cents


def first_float(data: dict, keys: tuple[str, ...], default=None):
    for key in keys:
        if key in data and data[key] is not None:
            return as_float(data[key])
    return default


def is_resolved_position(position: dict) -> bool:
    cur_price = position.get("curPrice")
    return bool(position.get("redeemable", False)) or (
        cur_price is not None and as_float(cur_price) <= 0
    )


def row_pl_components(position: dict) -> tuple[float, float]:
    """
    Return (realized_pl, unrealized_pl) for one API position.

    The current positions endpoint usually exposes cashPnl for open positions.
    If an upstream payload provides explicit realized/unrealized fields, prefer
    them. If a resolved row only has cashPnl, treat it as realized P/L.
    """
    cash_pnl = as_float(position.get("cashPnl"))
    realized = first_float(
        position,
        ("realizedPnl", "realizedPnL", "realizedPL", "realized_pl"),
    )
    unrealized = first_float(
        position,
        ("unrealizedPnl", "unrealizedPnL", "unrealizedPL", "unrealized_pl"),
    )

    if realized is None and unrealized is None:
        if is_resolved_position(position):
            return cash_pnl, 0.0
        return 0.0, cash_pnl
    return realized or 0.0, unrealized or 0.0


def build_position_row(
    addr: str,
    label: str,
    position: dict,
    category: str,
    median: float,
    port_total: float,
) -> dict:
    initial = as_float(position.get("initialValue"))
    cur_val = as_float(position.get("currentValue"))
    shares = as_float(position.get("size") or position.get("shares"))
    realized_pl, unrealized_pl = row_pl_components(position)
    row_pl = realized_pl + unrealized_pl
    row_roi_pct = row_pl / initial * 100 if initial > 0 else 0.0
    pct_port = round(cur_val / port_total * 100, 1) if port_total > 0 else 0.0
    avg_price = round(as_float(position.get("avgPrice")) * 100, 1)
    cur_price = round(as_float(position.get("curPrice")) * 100, 1)
    if shares <= 0 and avg_price > 0:
        shares = initial / (avg_price / 100)
    resolved = is_resolved_position(position)
    win = 1 if resolved and (realized_pl > 0 or bool(position.get("redeemable", False))) else 0
    loss = 1 if resolved and not win and (
        realized_pl < 0 or as_float(position.get("curPrice"), 1.0) <= 0
    ) else 0

    condition_id = position.get("conditionId") or position.get("condition_id")
    asset_id = position.get("asset") or position.get("assetId")
    event_id = position.get("eventId") or position.get("event_id") or position.get("eventSlug")
    event_slug = position.get("eventSlug") or position.get("event_slug") or ""
    market_id = condition_id or event_id or position.get("slug") or position.get("title", "?")
    source_link = f"https://polymarket.com/event/{event_slug}" if event_slug else ""

    return {
        "wallet": label,
        "addr": addr,
        "category": category,
        "market": position.get("title", "?"),
        "market_id": str(market_id),
        "condition_id": str(condition_id or ""),
        "asset_id": str(asset_id or ""),
        "event_id": str(event_id or ""),
        "event_slug": str(event_slug or ""),
        "row_id": (
            condition_id
            or asset_id
            or position.get("slug")
            or position.get("title", "?")
        ),
        "outcome": position.get("outcome", "?"),
        "side": position.get("outcome", "?"),
        "end_date": (position.get("endDate") or "")[:10],
        "resolution_time": position.get("endDate") or "",
        "last_activity_ts": position.get("lastTradePriceTimestamp") or position.get("timestamp") or "",
        "source_link": source_link,
        "avg_price": avg_price,
        "cur_price": cur_price,
        "avg_odds": american_odds_from_cents(avg_price),
        "cur_odds": american_odds_from_cents(cur_price),
        "avg_odds_value": american_odds_value_from_cents(avg_price),
        "cur_odds_value": american_odds_value_from_cents(cur_price),
        "size_usd": round(initial, 2),
        "shares": round(shares, 4),
        "current_value": round(cur_val, 2),
        "row_cost": round(initial, 2),
        "row_current_value": round(cur_val, 2),
        "row_realized_pl": round(realized_pl, 2),
        "row_unrealized_pl": round(unrealized_pl, 2),
        "row_pl": round(row_pl, 2),
        "row_roi_pct": round(row_roi_pct, 2),
        "row_win": win,
        "row_loss": loss,
        "pct_portfolio": pct_port,
        "conviction": conviction_score(initial, median),
        "tail_stake": round(MY_BASE_UNIT * conviction_score(initial, median) / 100, 2),
        "port_total": round(port_total, 2),
    }


def attach_category_aggregates(display_rows: list, category_source_rows: list) -> list:
    aggregates: dict[tuple[str, str], dict] = {}
    for row in category_source_rows:
        key = (row["addr"], row["category"])
        agg = aggregates.setdefault(
            key,
            {
                "category_total_cost": 0.0,
                "category_current_value": 0.0,
                "category_realized_pl": 0.0,
                "category_unrealized_pl": 0.0,
                "category_wins": 0,
                "category_losses": 0,
            },
        )
        agg["category_total_cost"] += row["row_cost"]
        agg["category_current_value"] += row["row_current_value"]
        agg["category_realized_pl"] += row["row_realized_pl"]
        agg["category_unrealized_pl"] += row["row_unrealized_pl"]
        agg["category_wins"] += row["row_win"]
        agg["category_losses"] += row["row_loss"]

    for agg in aggregates.values():
        agg["category_total_pl"] = agg["category_realized_pl"] + agg["category_unrealized_pl"]
        cost = agg["category_total_cost"]
        agg["category_roi_pct"] = agg["category_total_pl"] / cost * 100 if cost > 0 else 0.0

    for row in display_rows:
        agg = aggregates.get((row["addr"], row["category"]), {})
        wins = agg.get("category_wins", 0)
        losses = agg.get("category_losses", 0)
        decisions = wins + losses
        win_rate = wins / decisions * 100 if decisions > 0 else 0.0
        row.update({
            "category_total_cost": round(agg.get("category_total_cost", 0.0), 2),
            "category_current_value": round(agg.get("category_current_value", 0.0), 2),
            "category_realized_pl": round(agg.get("category_realized_pl", 0.0), 2),
            "category_unrealized_pl": round(agg.get("category_unrealized_pl", 0.0), 2),
            "category_total_pl": round(agg.get("category_total_pl", 0.0), 2),
            "category_roi_pct": round(agg.get("category_roi_pct", 0.0), 2),
            "category_wins": wins,
            "category_losses": losses,
            "category_wins_losses": f"{wins}-{losses}",
            "category_markets": decisions,
            "category_win_count": wins,
            "category_loss_count": losses,
            "category_win_rate_pct": round(win_rate, 2),
            "category_analytics_tag": row["category"],
            "category_analytics_source": "local",
        })
    return display_rows


def analytics_tag(category: str) -> str:
    return CATEGORY_ANALYTICS_TAGS.get(category, category)


def fetch_category_analytics(addr: str, category: str) -> Optional[dict]:
    if not ANALYTICS_STATS:
        return None

    tag = analytics_tag(category)
    try:
        r = requests.get(
            ANALYTICS_API,
            params={
                "tag": tag,
                "searchQuery": addr,
                "limit": 10,
                "offset": 0,
                "sortColumn": "rank",
                "sortDirection": "ASC",
            },
            timeout=15,
            headers={"User-Agent": "polymarket-tracker/1.0"},
        )
        if r.status_code != 200:
            print(f"[{addr[:8]}] analytics HTTP {r.status_code} for {tag}")
            return None

        payload = r.json()
        rows = payload.get("data") or []
        exact = [
            row for row in rows
            if str(row.get("trader", "")).lower() == addr.lower()
        ]
        row = exact[0] if exact else (rows[0] if rows else None)
        if not row:
            return None

        markets = int(as_float(row.get("event_ct")))
        wins = int(as_float(row.get("win_count")))
        losses = max(0, markets - wins)
        win_rate = as_float(row.get("win_rate")) * 100
        win_amount = as_float(row.get("win_amount"))
        loss_amount = abs(as_float(row.get("loss_amount")))
        current_value = as_float(row.get("total_current_value"))
        total_volume = win_amount + loss_amount + current_value
        total_pnl = as_float(row.get("overall_gain"))
        roi_pct = total_pnl / total_volume * 100 if total_volume > 0 else 0.0
        return {
            "category_analytics_tag": row.get("tag") or tag,
            "category_analytics_source": "polymarketanalytics",
            "category_markets": markets,
            "category_win_count": wins,
            "category_loss_count": losses,
            "category_win_rate_pct": round(win_rate, 2),
            "category_total_positions": int(as_float(row.get("total_positions"))),
            "category_active_positions": int(as_float(row.get("active_positions"))),
            "historical_category_total_pnl": round(total_pnl, 2),
            "historical_category_roi_pct": round(roi_pct, 2),
            "historical_category_total_volume": round(total_volume, 2),
            "historical_category_resolved_volume": round(win_amount + loss_amount, 2),
            "historical_category_open_volume": round(current_value, 2),
            "historical_source_url": (
                "https://legacy.polymarketanalytics.com/traders"
                f"?overallCategory={tag.replace(' ', '+')}&search={addr}"
            ),
        }
    except Exception as exc:
        print(f"[{addr[:8]}] analytics fetch error for {tag}: {exc}")
        return None


def fetch_category_analytics_map(rows: list) -> dict[tuple[str, str], dict]:
    keys = sorted({(row["addr"], row["category"]) for row in rows})
    if not keys:
        return {}

    result = {}
    with ThreadPoolExecutor(max_workers=min(8, len(keys))) as ex:
        futures = {
            ex.submit(fetch_category_analytics, addr, category): (addr, category)
            for addr, category in keys
        }
        for future in as_completed(futures):
            key = futures[future]
            stats = future.result()
            if stats:
                result[key] = stats
    return result


def attach_analytics_stats(display_rows: list, analytics: dict[tuple[str, str], dict]) -> list:
    for row in display_rows:
        stats = analytics.get((row["addr"], row["category"]))
        if not stats:
            continue

        row.update(stats)
        if "historical_category_total_pnl" in stats:
            row["category_total_pl"] = stats["historical_category_total_pnl"]
            row["category_roi_pct"] = stats["historical_category_roi_pct"]
            row["category_current_value"] = stats["historical_category_open_volume"]
        row["category_wins"] = stats["category_win_count"]
        row["category_losses"] = stats["category_loss_count"]
        row["category_wins_losses"] = (
            f"{stats['category_win_count']}-{stats['category_loss_count']}"
        )
    return display_rows


def attach_selected_wallet_stats(display_rows: list) -> list:
    for row in display_rows:
        meta = selected_wallet_meta(row["addr"], row["category"])
        if not meta:
            continue

        row.update({
            "historical_category_total_pnl": as_float(meta.get("total_pnl")),
            "historical_category_roi_pct": as_float(meta.get("roi_pct")),
            "historical_category_total_volume": as_float(meta.get("total_volume")),
            "historical_category_resolved_volume": as_float(meta.get("resolved_volume")),
            "historical_category_open_volume": as_float(meta.get("open_volume")),
            "historical_source_url": meta.get("source_url") or row.get("historical_source_url", ""),
            "category_win_rate_pct": as_float(meta.get("win_rate"), row.get("category_win_rate_pct", 0)),
            "category_total_positions": int(as_float(meta.get("number_of_trades"))),
            "category_active_positions": int(as_float(meta.get("number_of_open_positions"))),
            "category_markets": int(as_float(meta.get("number_of_resolved_positions"))),
            "category_win_count": int(as_float(meta.get("win_count"))),
            "category_analytics_source": "sharp_wallets_selected",
        })
        row["category_total_pl"] = round(row["historical_category_total_pnl"], 2)
        row["category_roi_pct"] = round(row["historical_category_roi_pct"], 2)
        row["category_current_value"] = round(row["historical_category_open_volume"], 2)
        losses = max(0, row["category_markets"] - row["category_win_count"])
        row["category_loss_count"] = losses
        row["category_wins"] = row["category_win_count"]
        row["category_losses"] = losses
        row["category_wins_losses"] = f"{row['category_win_count']}-{losses}"
    return display_rows


def position_key(row: dict) -> tuple:
    return (str(row.get("row_id", "")), row.get("outcome", "?"))


def is_sharp_for_category(addr: str, category: str) -> bool:
    addr = (addr or "").lower()
    if addr in MANUAL_WALLETS:
        return True
    return addr in SHARP_WALLET_POOLS.get(category, set())


def attach_sharp_wallet_counts(display_rows: list, source_rows: list) -> list:
    holders: dict[tuple, dict] = {}
    for row in source_rows:
        if not is_sharp_for_category(row["addr"], row["category"]):
            continue
        key = position_key(row)
        holder = holders.setdefault(key, {})
        holder[row["addr"]] = row["wallet"]

    for row in display_rows:
        holder = holders.get(position_key(row), {})
        row["sharp_wallet_count"] = len(holder)
        row["other_sharp_wallet_count"] = max(0, len(holder) - 1)
        row["sharp_wallets"] = ", ".join(sorted(holder.values()))
    return display_rows


def same_market(row: dict, market_id: str, category: str = None, resolution_time: str = None) -> bool:
    if str(row.get("market_id", "")) != str(market_id):
        return False
    if category and row.get("category") != category:
        return False
    if resolution_time and row.get("resolution_time") and row.get("resolution_time") != resolution_time:
        return False
    return True


def aggregate_wallet_position(rows: list) -> dict:
    first = rows[0]
    total_cost = sum(as_float(r.get("row_cost")) for r in rows)
    total_value = sum(as_float(r.get("row_current_value")) for r in rows)
    total_shares = sum(as_float(r.get("shares")) for r in rows)
    realized_pl = sum(as_float(r.get("row_realized_pl")) for r in rows)
    unrealized_pl = sum(as_float(r.get("row_unrealized_pl")) for r in rows)
    total_pl = realized_pl + unrealized_pl
    roi_pct = total_pl / total_cost * 100 if total_cost > 0 else 0.0
    weighted_entry = (
        sum(as_float(r.get("avg_price")) * as_float(r.get("row_cost")) for r in rows) / total_cost
        if total_cost > 0 else as_float(first.get("avg_price"))
    )
    weighted_current = (
        sum(as_float(r.get("cur_price")) * as_float(r.get("row_current_value")) for r in rows) / total_value
        if total_value > 0 else as_float(first.get("cur_price"))
    )
    category_pl = first.get("historical_category_total_pnl", first.get("category_total_pl", 0))
    category_roi = first.get("historical_category_roi_pct", first.get("category_roi_pct", 0))
    category_value = first.get("historical_category_open_volume", first.get("category_current_value", 0))
    category_total_volume = first.get("historical_category_total_volume", first.get("category_total_cost", 0))
    source_link = first.get("historical_source_url") or first.get("source_link", "")
    wallet_avg_size = as_float(first.get("wallet_avg_position_size")) or total_cost
    category_avg_size = as_float(first.get("category_avg_position_size")) or total_cost

    return {
        "wallet": first.get("wallet"),
        "addr": first.get("addr"),
        "side": first.get("side") or first.get("outcome"),
        "outcome": first.get("outcome"),
        "entry_price": round(weighted_entry, 2),
        "entry_odds": american_odds_from_cents(weighted_entry),
        "current_price": round(weighted_current, 2),
        "current_odds": american_odds_from_cents(weighted_current),
        "position_size": round(total_cost, 2),
        "shares": round(total_shares, 4),
        "cost_basis": round(total_cost, 2),
        "current_value": round(total_value, 2),
        "unrealized_pl": round(unrealized_pl, 2),
        "realized_pl": round(realized_pl, 2),
        "total_pl": round(total_pl, 2),
        "roi_pct": round(roi_pct, 2),
        "portfolio_size": first.get("port_total", 0),
        "wallet_total_portfolio_value": first.get("port_total", 0),
        "portfolio_pct": round(total_value / as_float(first.get("port_total")) * 100, 2) if as_float(first.get("port_total")) > 0 else 0.0,
        "wallet_avg_position_size": round(wallet_avg_size, 2),
        "position_size_multiple": round(total_cost / wallet_avg_size, 2) if wallet_avg_size > 0 else 0.0,
        "category_avg_position_size": round(category_avg_size, 2),
        "category_position_size_multiple": round(total_cost / category_avg_size, 2) if category_avg_size > 0 else 0.0,
        "wallet_category_portfolio_value": round(as_float(category_value), 2),
        "wallet_category_pl": round(as_float(category_pl), 2),
        "wallet_category_roi_pct": round(as_float(category_roi), 2),
        "wallet_category_total_volume": round(as_float(category_total_volume), 2),
        "wallet_category_resolved_volume": round(as_float(first.get("historical_category_resolved_volume")), 2),
        "wallet_historical_win_rate": first.get("category_win_rate_pct", 0),
        "wallet_historical_roi": round(as_float(category_roi), 2),
        "wallet_category_specific_roi": round(as_float(category_roi), 2),
        "wallet_category_stats_source": first.get("category_analytics_source", "local"),
        "position_conviction": max(as_float(r.get("conviction")) for r in rows),
        "wallet_conviction_contribution": max(as_float(r.get("sharp_consensus_score")) for r in rows),
        "sharp_for_category": as_float(first.get("category_markets")) >= 25 and as_float(first.get("category_win_rate_pct")) >= 52,
        "last_activity_timestamp": max(str(r.get("last_activity_ts") or "") for r in rows),
        "source_link": source_link,
        "duplicate_fill_count": len(rows),
    }


def get_position_wallet_details(
    rows: list,
    market_id: str,
    outcome_side: str,
    category: str = None,
    resolution_time: str = None,
) -> dict:
    market_rows = [
        row for row in rows
        if same_market(row, market_id, category=category, resolution_time=resolution_time)
    ]
    grouped: dict[tuple[str, str], list] = {}
    for row in market_rows:
        key = (row.get("addr"), row.get("side") or row.get("outcome"))
        grouped.setdefault(key, []).append(row)

    wallet_rows = [aggregate_wallet_position(items) for items in grouped.values()]
    aligned = [row for row in wallet_rows if row["side"] == outcome_side]
    opposing = [row for row in wallet_rows if row["side"] != outcome_side]
    aligned.sort(key=lambda row: row["position_size"], reverse=True)
    opposing.sort(key=lambda row: row["position_size"], reverse=True)

    selected = next(
        (row for row in market_rows if (row.get("side") or row.get("outcome")) == outcome_side),
        market_rows[0] if market_rows else {},
    )
    aligned_exposure = sum(as_float(row["position_size"]) for row in aligned)
    opposing_exposure = sum(as_float(row["position_size"]) for row in opposing)
    aligned_roi = statistics.mean([as_float(row["roi_pct"]) for row in aligned]) if aligned else 0.0
    aligned_category_roi = (
        statistics.mean([as_float(row["wallet_category_roi_pct"]) for row in aligned])
        if aligned else 0.0
    )
    final_conviction = max([as_float(row["position_conviction"]) for row in aligned], default=0.0)

    return {
        "summary": {
            "market_id": str(market_id),
            "market_title": selected.get("market", ""),
            "category": selected.get("category", category or ""),
            "selected_side": outcome_side,
            "resolution_time": selected.get("resolution_time", resolution_time or ""),
            "aligned_sharp_wallet_count": len(aligned),
            "aligned_wallet_count": len(aligned),
            "total_aligned_exposure": round(aligned_exposure, 2),
            "average_aligned_wallet_roi": round(aligned_roi, 2),
            "average_aligned_category_roi": round(aligned_category_roi, 2),
            "opposing_sharp_wallet_count": len(opposing),
            "opposing_wallet_count": len(opposing),
            "opposing_exposure": round(opposing_exposure, 2),
            "net_sharp_alignment": len(aligned) - len(opposing),
            "net_sharp_exposure": round(aligned_exposure - opposing_exposure, 2),
            "final_conviction_score": round(final_conviction),
        },
        "aligned_wallets": aligned,
        "opposing_wallets": opposing,
    }


getPositionWalletDetails = get_position_wallet_details


def attach_position_details(display_rows: list, source_rows: list) -> list:
    cache = {}
    for row in display_rows:
        key = (
            row.get("market_id"),
            row.get("side") or row.get("outcome"),
            row.get("category"),
            row.get("resolution_time"),
        )
        if key not in cache:
            cache[key] = get_position_wallet_details(
                source_rows,
                key[0],
                key[1],
                category=key[2],
                resolution_time=key[3],
            )
        row["position_details"] = cache[key]
        summary = cache[key]["summary"]
        row["aligned_sharp_wallet_count"] = summary.get("aligned_sharp_wallet_count", 0)
        row["opposing_sharp_wallet_count"] = summary.get("opposing_sharp_wallet_count", 0)
        row["net_sharp_alignment"] = summary.get("net_sharp_alignment", 0)
        row["net_sharp_exposure"] = summary.get("net_sharp_exposure", 0)
        final_score = summary.get("final_conviction_score", row.get("conviction", 0))
        row["conviction"] = final_score
        row["tail_stake"] = round(MY_BASE_UNIT * as_float(final_score) / 100, 2)
    return display_rows


def position_card_key(row: dict) -> tuple:
    return (
        str(row.get("market_id", "")),
        str(row.get("event_id", "")),
        row.get("side") or row.get("outcome"),
        row.get("category"),
        row.get("resolution_time") or row.get("end_date") or "",
    )


def collapse_position_cards(rows: list) -> list:
    """
    Collapse wallet-level display rows into one dashboard card per market side.

    The drawer still contains one aggregated wallet-position row per wallet.
    Card-level money fields are aligned-wallet totals for the selected side.
    """
    grouped: dict[tuple, list] = {}
    for row in rows:
        grouped.setdefault(position_card_key(row), []).append(row)

    cards = []
    for group_rows in grouped.values():
        representative = max(group_rows, key=lambda r: as_float(r.get("current_value")))
        card = representative.copy()
        details = card.get("position_details") or {}
        summary = details.get("summary", {})
        aligned = details.get("aligned_wallets", [])

        wallet_names = sorted({wallet.get("wallet") for wallet in aligned if wallet.get("wallet")})
        wallet_addrs = sorted({wallet.get("addr") for wallet in aligned if wallet.get("addr")})
        aligned_cost = sum(as_float(wallet.get("cost_basis")) for wallet in aligned)
        aligned_value = sum(as_float(wallet.get("current_value")) for wallet in aligned)
        aligned_realized = sum(as_float(wallet.get("realized_pl")) for wallet in aligned)
        aligned_unrealized = sum(as_float(wallet.get("unrealized_pl")) for wallet in aligned)
        aligned_pl = aligned_realized + aligned_unrealized
        aligned_portfolio = sum(as_float(wallet.get("wallet_total_portfolio_value")) for wallet in aligned)
        aligned_category_volume = sum(as_float(wallet.get("wallet_category_total_volume")) for wallet in aligned)
        aligned_category_value = sum(as_float(wallet.get("wallet_category_portfolio_value")) for wallet in aligned)
        aligned_category_pl = sum(as_float(wallet.get("wallet_category_pl")) for wallet in aligned)
        category_roi_samples = [
            as_float(wallet.get("wallet_category_roi_pct"))
            for wallet in aligned
            if wallet.get("wallet_category_roi_pct") is not None
        ]
        win_rate_samples = [
            as_float(wallet.get("wallet_historical_win_rate"))
            for wallet in aligned
            if wallet.get("wallet_historical_win_rate") is not None
        ]

        if len(wallet_names) == 1:
            wallet_label = wallet_names[0]
        elif wallet_names:
            wallet_label = f"{len(wallet_names)} aligned"
        else:
            wallet_label = card.get("wallet", "")

        card.update({
            "position_card": True,
            "wallet": wallet_label,
            "wallets": wallet_names,
            "wallet_addresses": wallet_addrs,
            "sharp_wallets": ", ".join(wallet_names),
            "sharp_wallet_count": len(wallet_names),
            "other_sharp_wallet_count": max(0, len(wallet_names) - 1),
            "aligned_sharp_wallet_count": summary.get("aligned_sharp_wallet_count", len(wallet_names)),
            "opposing_sharp_wallet_count": summary.get("opposing_sharp_wallet_count", 0),
            "net_sharp_alignment": summary.get("net_sharp_alignment", 0),
            "net_sharp_exposure": summary.get("net_sharp_exposure", 0),
            "size_usd": round(aligned_cost, 2),
            "current_value": round(aligned_value, 2),
            "row_cost": round(aligned_cost, 2),
            "row_current_value": round(aligned_value, 2),
            "row_realized_pl": round(aligned_realized, 2),
            "row_unrealized_pl": round(aligned_unrealized, 2),
            "row_pl": round(aligned_pl, 2),
            "row_roi_pct": round(aligned_pl / aligned_cost * 100, 2) if aligned_cost > 0 else 0.0,
            "pct_portfolio": round(aligned_value / aligned_portfolio * 100, 2) if aligned_portfolio > 0 else 0.0,
            "category_total_cost": round(aligned_category_volume, 2),
            "category_current_value": round(aligned_category_value, 2),
            "category_total_pl": round(aligned_category_pl, 2),
            "category_roi_pct": (
                round(aligned_category_pl / aligned_category_volume * 100, 2)
                if aligned_category_volume > 0
                else round(statistics.mean(category_roi_samples), 2) if category_roi_samples else 0.0
            ),
            "category_win_rate_pct": (
                round(statistics.mean(win_rate_samples), 2) if win_rate_samples else card.get("category_win_rate_pct", 0.0)
            ),
            "conviction": summary.get("final_conviction_score", card.get("conviction", 0)),
        })
        card["tail_stake"] = round(MY_BASE_UNIT * as_float(card.get("conviction")) / 100, 2)
        cards.append(card)

    return cards


def _side_weighted_exposure(card: dict) -> float:
    """
    Conviction-weighted dollar signal for one side of a market.
    Each wallet contributes: position_size × size_multiple
    (a 3x average bet counts 3× more than a 0.5x average bet).
    Falls back to card-level fields when per-wallet detail is absent.
    """
    wallets = (card.get("position_details") or {}).get("aligned_wallets") or []
    if wallets:
        return sum(
            as_float(w.get("position_size"))
            * max(as_float(w.get("position_size_multiple") or 1), 0.1)
            for w in wallets
        )
    size = as_float(card.get("size_usd") or card.get("row_cost"))
    mult = max(as_float(card.get("position_size_multiple") or 1), 0.1)
    return size * mult


def dedup_by_market(cards: list) -> list:
    """
    One card per market event.

    When wallets are split across YES / NO, determine the dominant side by
    weighted exposure (Σ position_size × size_multiple per wallet), not head
    count.  A 3-wallet YES each betting 3× their average outweighs a 5-wallet
    NO each betting 0.5× their average — the YES side is not arbing/hedging,
    it's expressing real conviction.

    Conviction is then scaled by the weighted consensus ratio so a near-50/50
    split halves the score even when the dominant side 'wins' on exposure.
    """
    grouped: dict[tuple, list] = {}
    for card in cards:
        key = (
            str(card.get("market_id", "")),
            card.get("category", ""),
            card.get("resolution_time") or card.get("end_date") or "",
        )
        grouped.setdefault(key, []).append(card)

    result = []
    for siblings in grouped.values():
        if len(siblings) == 1:
            result.append(siblings[0])
            continue

        # Rank sides by weighted exposure
        siblings.sort(key=_side_weighted_exposure, reverse=True)
        dominant   = siblings[0]
        dom_we     = _side_weighted_exposure(dominant)
        other_we   = sum(_side_weighted_exposure(s) for s in siblings[1:])
        total_we   = dom_we + other_we

        # consensus ∈ [0, 1]: 1 = all weight on one side, 0 = perfect split
        consensus = (dom_we - other_we) / total_we if total_we > 0 else 1.0
        consensus = max(0.0, consensus)

        # Scale conviction: unanimous → unchanged, perfectly split → half score
        raw_conv = as_float(dominant.get("conviction"))
        adj_conv = max(1, round(raw_conv * (0.5 + 0.5 * consensus)))

        # Count opposing wallets for the UI alignment display
        opposing_wallets = sum(
            len((s.get("position_details") or {}).get("aligned_wallets") or [s])
            for s in siblings[1:]
        )
        dominant["conviction"]                  = adj_conv
        dominant["tail_stake"]                  = round(MY_BASE_UNIT * adj_conv / 100, 2)
        dominant["opposing_sharp_wallet_count"] = opposing_wallets
        dominant["net_sharp_alignment"]         = int(
            as_float(dominant.get("aligned_sharp_wallet_count")) - opposing_wallets
        )
        result.append(dominant)

    return result


def collapse_by_event(cards: list) -> list:
    """
    Fold market-level cards into one card per event slug.
    When a wallet has positions across multiple correlated markets in the same
    event (e.g. match-winner + BTTS + over/under), this collapses them into one
    aggregated card so hedged exposure isn't counted as separate positions.
    """
    grouped: dict[tuple, list] = {}
    for card in cards:
        event = card.get("event_slug") or card.get("event_id") or str(card.get("market_id", ""))
        key = (event, card.get("category", ""), card.get("resolution_time") or card.get("end_date") or "")
        grouped.setdefault(key, []).append(card)

    result = []
    for group_cards in grouped.values():
        if len(group_cards) == 1:
            result.append(group_cards[0])
            continue

        rep  = max(group_cards, key=lambda c: as_float(c.get("conviction")))
        card = rep.copy()

        total_cost       = sum(as_float(c.get("size_usd") or c.get("row_cost")) for c in group_cards)
        total_value      = sum(as_float(c.get("current_value") or c.get("row_current_value")) for c in group_cards)
        total_realized   = sum(as_float(c.get("row_realized_pl")) for c in group_cards)
        total_unrealized = sum(as_float(c.get("row_unrealized_pl")) for c in group_cards)
        total_pl         = total_realized + total_unrealized

        all_wallets: set = set()
        all_addrs: set   = set()
        for c in group_cards:
            all_wallets.update(c.get("wallets") or ([c.get("wallet")] if c.get("wallet") else []))
            all_addrs.update(c.get("wallet_addresses") or ([c.get("addr")] if c.get("addr") else []))

        wallet_label = rep.get("wallet", "")
        if len(all_wallets) > 1:
            wallet_label = f"{len(all_wallets)} aligned"

        event_slug  = rep.get("event_slug", "")
        event_title = " ".join(w.capitalize() for w in event_slug.split("-")) if event_slug else rep.get("market", "?")

        card.update({
            "market":             event_title,
            "outcome":            "Multi",
            "side":               "Multi",
            "event_market_count": len(group_cards),
            "sub_markets":        [c.get("market", "") for c in group_cards],
            "wallet":             wallet_label,
            "wallets":            sorted(all_wallets),
            "wallet_addresses":   sorted(all_addrs),
            "sharp_wallets":      ", ".join(sorted(all_wallets)),
            "sharp_wallet_count": len(all_wallets),
            "size_usd":           round(total_cost, 2),
            "current_value":      round(total_value, 2),
            "row_cost":           round(total_cost, 2),
            "row_current_value":  round(total_value, 2),
            "row_realized_pl":    round(total_realized, 2),
            "row_unrealized_pl":  round(total_unrealized, 2),
            "row_pl":             round(total_pl, 2),
            "row_roi_pct":        round(total_pl / total_cost * 100, 2) if total_cost > 0 else 0.0,
            "avg_price":          0,
            "cur_price":          0,
            "avg_odds":           "n/a",
            "cur_odds":           "n/a",
        })
        result.append(card)

    return result


def load_fill_history() -> dict:
    history = {
        "wallet_sizes": {},
        "category_sizes": {},
        "market_counts": {},
    }
    try:
        con = sqlite3.connect(DB)
        rows = con.execute(
            "SELECT wallet, category, condition_id, usd_size FROM fills"
        ).fetchall()
        con.close()
    except Exception:
        return history

    for wallet, category, condition_id, usd_size in rows:
        wallet = (wallet or "").lower()
        category = category or "Other"
        if not wallet:
            continue
        history["wallet_sizes"].setdefault(wallet, []).append(as_float(usd_size))
        history["category_sizes"].setdefault((wallet, category), []).append(as_float(usd_size))
        if condition_id:
            key = (wallet, category, str(condition_id))
            history["market_counts"][key] = history["market_counts"].get(key, 0) + 1
    return history


def add_current_size_fallbacks(history: dict, rows: list) -> dict:
    current_wallet_sizes = {}
    current_category_sizes = {}
    for row in rows:
        current_wallet_sizes.setdefault(row["addr"], []).append(row["row_cost"])
        current_category_sizes.setdefault((row["addr"], row["category"]), []).append(row["row_cost"])

    for key, values in current_wallet_sizes.items():
        if not history["wallet_sizes"].get(key):
            history["wallet_sizes"][key] = values
    for key, values in current_category_sizes.items():
        if not history["category_sizes"].get(key):
            history["category_sizes"][key] = values
    return history


def category_skill_score(row: dict) -> int:
    markets = as_float(row.get("category_markets"))
    win_rate = as_float(row.get("category_win_rate_pct"))
    if markets <= 0 or win_rate <= 0:
        return 50

    raw = clamp(50 + (win_rate - 50) * 2)
    confidence = min(1.0, (markets ** 0.5) / 10)
    return round(50 + (raw - 50) * confidence)


def concentration_score(row: dict) -> int:
    # Treat 10% of the wallet's open book as full concentration.
    return round(clamp(as_float(row.get("pct_portfolio")) / 10 * 100))


def activity_score(row: dict, history: dict) -> int:
    key = (row["addr"], row["category"], str(row.get("row_id", "")))
    fills = history["market_counts"].get(key, 0)
    if fills <= 0:
        return 50
    return round(clamp(40 + fills * 15))


def sharp_consensus_score(row: dict) -> int:
    # One wallet is neutral; each additional tracked sharp wallet adds signal.
    return round(clamp(50 + as_float(row.get("other_sharp_wallet_count")) * 20))


def attach_conviction_scores(display_rows: list, source_rows: list) -> list:
    history = add_current_size_fallbacks(load_fill_history(), source_rows)
    for row in display_rows:
        wallet_sizes = history["wallet_sizes"].get(row["addr"], [])
        category_sizes = history["category_sizes"].get((row["addr"], row["category"]), [])
        wallet_avg_size = statistics.mean(wallet_sizes) if wallet_sizes else row["row_cost"]
        category_avg_size = statistics.mean(category_sizes) if category_sizes else row["row_cost"]

        size_score = percentile_score(row["row_cost"], wallet_sizes)
        category_size_score = percentile_score(row["row_cost"], category_sizes)
        skill = category_skill_score(row)
        concentration = concentration_score(row)
        activity = activity_score(row, history)
        consensus = sharp_consensus_score(row)

        conviction = round(
            0.25 * size_score
            + 0.20 * category_size_score
            + 0.20 * skill
            + 0.15 * concentration
            + 0.10 * activity
            + 0.10 * consensus
        )

        row.update({
            "size_score": size_score,
            "category_size_score": category_size_score,
            "skill_score": skill,
            "concentration_score": concentration,
            "activity_score": activity,
            "sharp_consensus_score": consensus,
            "wallet_avg_position_size": round(wallet_avg_size, 2),
            "category_avg_position_size": round(category_avg_size, 2),
            "position_size_multiple": round(row["row_cost"] / wallet_avg_size, 2) if wallet_avg_size > 0 else 0.0,
            "category_position_size_multiple": round(row["row_cost"] / category_avg_size, 2) if category_avg_size > 0 else 0.0,
            "conviction": round(clamp(conviction)),
            "tail_stake": round(MY_BASE_UNIT * clamp(conviction) / 100, 2),
        })
    return display_rows


def fetch_positions(addr: str) -> list:
    try:
        r = requests.get(
            POSITIONS_API,
            params={"user": addr, "limit": 200},
            timeout=15,
            headers={"User-Agent": "polymarket-tracker/1.0"},
        )
        if r.status_code == 200:
            return r.json()
        print(f"[{addr[:8]}] positions HTTP {r.status_code}")
    except Exception as exc:
        print(f"[{addr[:8]}] positions fetch error: {exc}")
    return []


# ── Cache with background refresh ─────────────────────────────────────────────
_cache: dict = {"data": [], "updated": None, "error": None}
_lock = threading.Lock()


def build_positions() -> list:
    """
    Fetch all wallets in parallel, then apply three filters before returning:
      1. Sports only  — positions whose category is not "Other"
      2. Date window  — endDate is today through today + RESOLVE_WINDOW_DAYS
      3. (conviction filter is client-side via the slider)

    Portfolio % is computed against the wallet's TOTAL position value (all
    positions, before any filter) so the percentage is meaningful.

    Category P/L and ROI are computed by wallet + category across every sports
    position returned by the API, then joined back onto the visible rows.
    """
    try:
        con = sqlite3.connect(DB)
        medians = {addr: wallet_median(con, addr) for addr in WALLETS}
        con.close()
    except Exception:
        medians = {addr: 0.0 for addr in WALLETS}

    today   = datetime.now(timezone.utc).date()
    cutoff  = today + timedelta(days=RESOLVE_WINDOW)

    # Fetch all wallets in parallel
    raw: dict = {}
    with ThreadPoolExecutor(max_workers=len(WALLETS) or 1) as ex:
        futures = {ex.submit(fetch_positions, addr): addr for addr in WALLETS}
        for future in as_completed(futures):
            raw[futures[future]] = future.result()

    # If no fill history exists yet (monitor.py hasn't seeded seen.db), fall back
    # to the median initialValue across this wallet's current positions.
    # Less accurate than historical fills but gives real scores immediately.
    for addr, positions in raw.items():
        if medians.get(addr, 0.0) <= 0 and positions:
            sizes = [float(p.get("initialValue", 0) or 0) for p in positions if p.get("initialValue")]
            if sizes:
                medians[addr] = statistics.median(sizes)

    # Total portfolio value per wallet across ALL their positions (pre-filter),
    # so pct_portfolio reflects how much of their whole book this position is.
    portfolio_totals = {
        addr: sum(float(p.get("currentValue", 0) or 0) for p in positions)
        for addr, positions in raw.items()
    }

    category_source_rows = []
    result = []
    for addr, positions in raw.items():
        label        = WALLETS[addr]
        median       = medians.get(addr, 0.0)
        port_total   = portfolio_totals.get(addr, 0.0)

        for p in positions:
            title = p.get("title", "?")
            eslug = p.get("eventSlug", "") or ""

            # 1. Sports filter
            cat = classify_category(title, eslug)
            if cat == "Other":
                continue
            if not is_sharp_for_category(addr, cat):
                continue

            row = build_position_row(addr, label, p, cat, median, port_total)
            category_source_rows.append(row)

            # 2. Keep only genuinely open positions within the resolve window.
            #    Three conditions must ALL be true:
            #      a) endDate is today or within the next RESOLVE_WINDOW days
            #      b) curPrice > 0  — price hit 0 means the market resolved as a loss
            #      c) not redeemable — redeemable=True means it resolved as a win
            redeemable = bool(p.get("redeemable", False))
            cur_price  = as_float(p.get("curPrice"))
            end_str    = (p.get("endDate") or "")[:10]
            if redeemable:
                continue   # already resolved as a win — settled, skip
            if cur_price <= 0:
                continue   # price went to 0 — resolved as a loss, skip
            if not odds_in_range(row["cur_price"]):
                continue   # only show plays priced between -250 and +250
            if end_str:
                try:
                    end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
                    if end_date < today or end_date > cutoff:
                        continue   # outside the 0–2 day window
                except ValueError:
                    pass

            result.append(row.copy())

    analytics = fetch_category_analytics_map(category_source_rows)
    for rows in (category_source_rows, result):
        attach_category_aggregates(rows, category_source_rows)
        attach_analytics_stats(rows, analytics)
        attach_selected_wallet_stats(rows)
        attach_sharp_wallet_counts(rows, category_source_rows)
        attach_conviction_scores(rows, category_source_rows)
    attach_position_details(result, category_source_rows)
    result = collapse_position_cards(result)
    result = dedup_by_market(result)
    result = collapse_by_event(result)
    result.sort(key=lambda x: x["conviction"], reverse=True)
    return result


def refresh_loop():
    while True:
        try:
            data = build_positions()
            with _lock:
                _cache["data"]    = data
                _cache["updated"] = datetime.now(timezone.utc).strftime("%H:%M:%S UTC")
                _cache["error"]   = None
            print(f"[dashboard] refreshed — {len(data)} positions across {len(WALLETS)} wallets")
        except Exception as exc:
            with _lock:
                _cache["error"] = str(exc)
            print(f"[dashboard] refresh error: {exc}")
        time.sleep(REFRESH_SECONDS)


# ── Flask ──────────────────────────────────────────────────────────────────────
app = Flask(__name__)

from backtest_ui import backtest_bp  # noqa: E402
app.register_blueprint(backtest_bp)


@app.route("/api/positions")
def api_positions():
    with _lock:
        return jsonify({
            "data":    _cache["data"],
            "updated": _cache["updated"],
            "error":   _cache["error"],
        })


# ── Backtest / resolution API ──────────────────────────────────────────────────
# These routes read from backtest.db (written by snapshot.py / resolver.py).
# They return empty lists gracefully when no data has been collected yet.

def _backtest_conn():
    """Return a read-only connection to backtest.db, or None if it doesn't exist."""
    import os
    from backtest_db import get_conn, BACKTEST_DB
    if not os.path.exists(BACKTEST_DB):
        return None
    try:
        conn = get_conn(BACKTEST_DB)
        return conn
    except Exception:
        return None


@app.route("/api/backtest/summary")
def api_backtest_summary():
    """
    Wallet performance summaries from wallet_performance_summary.
    Query params: wallet (label/addr substring), category
    """
    from flask import request as req
    wallet_q   = req.args.get("wallet", "")
    category_q = req.args.get("category", "")

    conn = _backtest_conn()
    if conn is None:
        return jsonify({"data": [], "note": "backtest.db not found — run snapshot.py first"})

    wheres: list[str] = []
    params: list      = []
    if wallet_q:
        wheres.append("(wps.wallet_address LIKE ? OR tw.wallet_label LIKE ?)")
        params.extend([f"%{wallet_q}%", f"%{wallet_q}%"])
    if category_q:
        wheres.append("wps.category = ?")
        params.append(category_q)

    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""

    rows = conn.execute(f"""
        SELECT
            wps.*,
            COALESCE(tw.wallet_label, wps.wallet_address) AS label
        FROM wallet_performance_summary wps
        LEFT JOIN tracked_wallets tw ON tw.wallet_address = wps.wallet_address
        {where_sql}
        ORDER BY wps.roi_pct DESC
    """, params).fetchall()

    conn.close()
    return jsonify({"data": [dict(r) for r in rows]})


@app.route("/api/backtest/resolved")
def api_backtest_resolved():
    """
    Recently resolved positions.
    Query params: limit (default 50), wallet, category, win_loss (WIN|LOSS|PUSH|NEEDS_REVIEW)
    """
    from flask import request as req
    limit    = min(int(req.args.get("limit", 50)), 500)
    wallet_q = req.args.get("wallet", "")
    cat_q    = req.args.get("category", "")
    wl_q     = req.args.get("win_loss", "")

    conn = _backtest_conn()
    if conn is None:
        return jsonify({"data": [], "note": "backtest.db not found — run snapshot.py first"})

    wheres: list[str] = []
    params: list      = []
    if wallet_q:
        wheres.append("(wallet_address LIKE ? OR wallet_label LIKE ?)")
        params.extend([f"%{wallet_q}%", f"%{wallet_q}%"])
    if cat_q:
        wheres.append("category = ?")
        params.append(cat_q)
    if wl_q:
        wheres.append("win_loss = ?")
        params.append(wl_q.upper())

    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""

    rows = conn.execute(f"""
        SELECT * FROM resolved_positions
        {where_sql}
        ORDER BY resolved_at DESC
        LIMIT ?
    """, [*params, limit]).fetchall()

    conn.close()
    return jsonify({"data": [dict(r) for r in rows]})


@app.route("/api/backtest/snapshots")
def api_backtest_snapshots():
    """
    Latest position snapshot per (wallet, token).
    Query params: wallet, category, limit (default 100)
    """
    from flask import request as req
    limit    = min(int(req.args.get("limit", 100)), 1000)
    wallet_q = req.args.get("wallet", "")
    cat_q    = req.args.get("category", "")

    conn = _backtest_conn()
    if conn is None:
        return jsonify({"data": [], "note": "backtest.db not found — run snapshot.py first"})

    wheres: list[str] = []
    params: list      = []
    if wallet_q:
        wheres.append("(wallet_address LIKE ? OR wallet_label LIKE ?)")
        params.extend([f"%{wallet_q}%", f"%{wallet_q}%"])
    if cat_q:
        wheres.append("category = ?")
        params.append(cat_q)

    where_sql = ("WHERE " + " AND ".join(wheres)) if wheres else ""

    rows = conn.execute(f"""
        SELECT ps.*
        FROM position_snapshots ps
        INNER JOIN (
            SELECT wallet_address, token_id, MAX(snapshot_ts) AS latest
            FROM position_snapshots
            GROUP BY wallet_address, token_id
        ) latest ON latest.wallet_address = ps.wallet_address
               AND latest.token_id        = ps.token_id
               AND latest.latest          = ps.snapshot_ts
        {where_sql}
        ORDER BY ps.cost_basis DESC
        LIMIT ?
    """, [*params, limit]).fetchall()

    conn.close()
    return jsonify({"data": [dict(r) for r in rows]})


@app.route("/export")
def export_excel():
    import io
    from flask import send_file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "openpyxl not installed — run: pip install openpyxl", 500

    with _lock:
        positions = list(_cache.get("data") or [])

    # ── style helpers ──────────────────────────────────────────────────────────
    def fill(hex6):
        return PatternFill("solid", fgColor="FF" + hex6.lstrip("#").upper())

    def fnt(hex6, bold=False, size=10):
        return Font(color="FF" + hex6.lstrip("#").upper(), bold=bold, size=size)

    HDR_FILL = fill("0D1F16")
    HDR_FONT = fnt("34C759", bold=True, size=10)
    YES_FILL = fill("0A1F12")
    NO_FILL  = fill("1F0D0D")
    MLT_FILL = fill("1F1700")
    DEF_FILL = fill("111814")
    POS_FONT = fnt("34C759", bold=True)
    NEG_FONT = fnt("F0606E", bold=True)
    GLD_FONT = fnt("E6A817", bold=True)
    WHT_FONT = fnt("D8EBE0")
    DIM_FONT = fnt("5A7A68")
    CENTER   = Alignment(horizontal="center", vertical="center")
    LEFT     = Alignment(horizontal="left", vertical="center", wrap_text=False)

    def hdr_row(ws, headers):
        for col, (title, width) in enumerate(headers, 1):
            c = ws.cell(1, col, title)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    def sc(ws, row, col, val, f=None, fn=None, al=None, fmt=None):
        c = ws.cell(row, col, val)
        if f:   c.fill            = f
        if fn:  c.font            = fn
        if al:  c.alignment       = al
        if fmt: c.number_format   = fmt
        return c

    wb = Workbook()

    # ── Sheet 1: Open Positions ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Open Positions"
    hdr_row(ws1, [
        ("Conv", 6), ("Sport", 11), ("Market", 46), ("Side", 6),
        ("Entry Odds", 11), ("Cur Odds", 9), ("Invested $", 11),
        ("Cur Value $", 11), ("P&L $", 10), ("ROI %", 8),
        ("Tail $", 8), ("Aligned", 8), ("Sharp Wallets", 28), ("Ends", 12),
    ])
    for ri, pos in enumerate(positions, 2):
        outcome  = pos.get("outcome", "")
        rf       = YES_FILL if outcome == "Yes" else (NO_FILL if outcome == "No" else MLT_FILL)
        pl       = as_float(pos.get("row_pl", 0))
        roi      = as_float(pos.get("row_roi_pct", 0))
        sf       = (fnt("34C759", bold=True) if outcome == "Yes"
                    else fnt("F0606E", bold=True) if outcome == "No" else GLD_FONT)
        ends     = (pos.get("end_date") or pos.get("resolution_time") or "")[:10]
        for ci, (v, fn, al, fmt) in enumerate([
            (pos.get("conviction", 0),              fnt("E0ECE4", bold=True), CENTER, None),
            (pos.get("category", ""),               DIM_FONT, CENTER, None),
            (pos.get("market", ""),                 WHT_FONT, LEFT,   None),
            (outcome,                               sf,       CENTER, None),
            (pos.get("avg_odds", ""),               DIM_FONT, CENTER, None),
            (pos.get("cur_odds", ""),               WHT_FONT, CENTER, None),
            (as_float(pos.get("size_usd", 0)),      DIM_FONT, CENTER, "#,##0.00"),
            (as_float(pos.get("current_value", 0)), WHT_FONT, CENTER, "#,##0.00"),
            (pl,  POS_FONT if pl  > 0 else NEG_FONT if pl  < 0 else WHT_FONT, CENTER, "#,##0.00"),
            (roi, POS_FONT if roi > 0 else NEG_FONT if roi < 0 else WHT_FONT, CENTER, "0.00"),
            (as_float(pos.get("tail_stake", 0)),    DIM_FONT, CENTER, "0.00"),
            (pos.get("sharp_wallet_count", 0),      DIM_FONT, CENTER, None),
            (pos.get("sharp_wallets", ""),          DIM_FONT, LEFT,   None),
            (ends,                                  DIM_FONT, CENTER, None),
        ], 1):
            sc(ws1, ri, ci, v, f=rf, fn=fn, al=al, fmt=fmt)
        ws1.row_dimensions[ri].height = 17

    # ── Sheet 2: Sharp Wallet DB ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Sharp Wallet DB")
    hdr_row(ws2, [
        ("Label", 16), ("Address", 44), ("Sport", 12),
        ("Sharpness", 11), ("Win Rate %", 11), ("Markets", 10),
        ("Wins", 8), ("Volume $", 14), ("P&L $", 14),
        ("ROI %", 8), ("Rank", 6),
    ])
    sharp_data = []
    if os.path.exists(SHARP_WALLETS_FILE):
        try:
            with open(SHARP_WALLETS_FILE, "r", encoding="utf-8") as f_:
                sharp_data = json.load(f_).get("wallets", [])
        except Exception:
            pass
    for ri, w in enumerate(sharp_data, 2):
        wr      = as_float(w.get("win_rate", 0))
        roi_v   = as_float(w.get("roi_pct", 0))
        pl_v    = as_float(w.get("total_pnl", 0))
        sharp_v = as_float(w.get("sharpness_score", 0))
        for ci, (v, fn, al, fmt) in enumerate([
            (w.get("wallet_label", ""),                     WHT_FONT, LEFT,   None),
            (w.get("wallet_address", ""),                   DIM_FONT, LEFT,   None),
            (w.get("category", ""),                         DIM_FONT, CENTER, None),
            (sharp_v, GLD_FONT if sharp_v >= 70 else WHT_FONT, CENTER, "0.0"),
            (wr, fnt("34C759", bold=True) if wr >= 55 else fnt("F0606E", bold=True) if wr < 50 else WHT_FONT, CENTER, "0.00"),
            (int(as_float(w.get("number_of_resolved_positions", 0))), DIM_FONT, CENTER, None),
            (int(as_float(w.get("win_count", 0))),          fnt("34C759"), CENTER, None),
            (as_float(w.get("total_volume", 0)),            DIM_FONT, CENTER, "#,##0"),
            (pl_v, POS_FONT if pl_v > 0 else NEG_FONT,     CENTER, "#,##0.00"),
            (roi_v, POS_FONT if roi_v > 0 else NEG_FONT if roi_v < 0 else WHT_FONT, CENTER, "0.00"),
            (int(as_float(w.get("rank", 0))),               DIM_FONT, CENTER, None),
        ], 1):
            sc(ws2, ri, ci, v, f=DEF_FILL, fn=fn, al=al, fmt=fmt)
        ws2.row_dimensions[ri].height = 16

    # ── Sheet 3: Category Summary ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Category Summary")
    hdr_row(ws3, [
        ("Sport", 14), ("Positions", 10), ("Wallets", 9),
        ("Invested $", 12), ("Cur Value $", 12), ("P&L $", 12),
        ("ROI %", 8), ("Avg Win %", 10),
    ])
    by_cat: dict = {}
    for pos in positions:
        cat = pos.get("category", "Other")
        g   = by_cat.setdefault(cat, {"count": 0, "wallets": set(), "invested": 0.0, "value": 0.0, "pl": 0.0, "wrs": []})
        g["count"]    += 1
        for wn in (pos.get("wallets") or [pos.get("wallet", "")]):
            if wn: g["wallets"].add(wn)
        g["invested"] += as_float(pos.get("size_usd", 0))
        g["value"]    += as_float(pos.get("current_value", 0))
        g["pl"]       += as_float(pos.get("row_pl", 0))
        wr_ = pos.get("category_win_rate_pct")
        if wr_ is not None: g["wrs"].append(as_float(wr_))
    for ri, (cat, g) in enumerate(sorted(by_cat.items(), key=lambda x: -x[1]["pl"]), 2):
        pl_v  = g["pl"]
        roi_v = pl_v / g["invested"] * 100 if g["invested"] > 0 else 0.0
        wr_v  = sum(g["wrs"]) / len(g["wrs"]) if g["wrs"] else 0.0
        for ci, (v, fn, al, fmt) in enumerate([
            (cat,                WHT_FONT, LEFT,   None),
            (g["count"],         DIM_FONT, CENTER, None),
            (len(g["wallets"]), DIM_FONT,  CENTER, None),
            (g["invested"],      DIM_FONT, CENTER, "#,##0.00"),
            (g["value"],         DIM_FONT, CENTER, "#,##0.00"),
            (pl_v,  POS_FONT if pl_v  > 0 else NEG_FONT if pl_v  < 0 else WHT_FONT, CENTER, "#,##0.00"),
            (roi_v, POS_FONT if roi_v > 0 else NEG_FONT if roi_v < 0 else WHT_FONT, CENTER, "0.00"),
            (wr_v,               DIM_FONT, CENTER, "0.0"),
        ], 1):
            sc(ws3, ri, ci, v, f=DEF_FILL, fn=fn, al=al, fmt=fmt)
        ws3.row_dimensions[ri].height = 16

    # ── Sheet 4: Wallet Summary ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Wallet Summary")
    hdr_row(ws4, [
        ("Wallet", 16), ("Positions", 10), ("Invested $", 12),
        ("Cur Value $", 12), ("P&L $", 12), ("ROI %", 8),
        ("Avg Conv", 9), ("Sports", 30),
    ])
    by_wlt: dict = {}
    for pos in positions:
        for wn in (pos.get("wallets") or [pos.get("wallet", "?")]):
            g = by_wlt.setdefault(wn, {"count": 0, "invested": 0.0, "value": 0.0, "pl": 0.0, "convs": [], "cats": set()})
            g["count"]    += 1
            g["invested"] += as_float(pos.get("size_usd", 0))
            g["value"]    += as_float(pos.get("current_value", 0))
            g["pl"]       += as_float(pos.get("row_pl", 0))
            g["convs"].append(as_float(pos.get("conviction", 0)))
            cat = pos.get("category")
            if cat: g["cats"].add(cat)
    for ri, (wn, g) in enumerate(sorted(by_wlt.items(), key=lambda x: -x[1]["pl"]), 2):
        pl_v  = g["pl"]
        roi_v = pl_v / g["invested"] * 100 if g["invested"] > 0 else 0.0
        cv_v  = sum(g["convs"]) / len(g["convs"]) if g["convs"] else 0.0
        for ci, (v, fn, al, fmt) in enumerate([
            (wn,                   WHT_FONT, LEFT,   None),
            (g["count"],           DIM_FONT, CENTER, None),
            (g["invested"],        DIM_FONT, CENTER, "#,##0.00"),
            (g["value"],           DIM_FONT, CENTER, "#,##0.00"),
            (pl_v,  POS_FONT if pl_v  > 0 else NEG_FONT if pl_v  < 0 else WHT_FONT, CENTER, "#,##0.00"),
            (roi_v, POS_FONT if roi_v > 0 else NEG_FONT if roi_v < 0 else WHT_FONT, CENTER, "0.00"),
            (round(cv_v, 1),       GLD_FONT, CENTER, "0.0"),
            (", ".join(sorted(g["cats"])), DIM_FONT, LEFT, None),
        ], 1):
            sc(ws4, ri, ci, v, f=DEF_FILL, fn=fn, al=al, fmt=fmt)
        ws4.row_dimensions[ri].height = 16

    # ── stream ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"polytrack_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/")
def index():
    return render_template_string(HTML, refresh=REFRESH_SECONDS)


# ── HTML template ──────────────────────────────────────────────────────────────
HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Polymarket Tracker</title>
<style>
  /*
   * Palette
   * --bg:       #0e1412  main background
   * --surface:  #151c18  card / raised surfaces
   * --border:   #222e27  all borders
   * --accent:   #34c759  emerald accent (not neon)
   * --red:      #f0606e  soft red
   * --gold:     #e6a817  muted gold
   * --t1:       #e0ece4  primary text
   * --t2:       #8aa898  secondary text
   * --t3:       #4d6659  muted text
   */
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: #0e1412; color: #96b8a2; font-family: 'Segoe UI', system-ui, sans-serif; font-size: 13px; min-height: 100vh; }
  a { color: inherit; text-decoration: none; }

  /* ── Chrome ───────────────────────────────────────────────── */
  header { padding: 12px 20px; border-bottom: 1px solid #222e27; display: flex; align-items: center; gap: 14px; flex-wrap: wrap; background: #0e1412; }
  h1 { font-size: 15px; font-weight: 700; color: #34c759; letter-spacing: .2px; }
  .badge { background: #151c18; border: 1px solid #222e27; border-radius: 10px; padding: 2px 9px; font-size: 11px; color: #4d6659; }
  #meta { margin-left: auto; color: #4d6659; font-size: 11px; }
  #spinner { width: 9px; height: 9px; border: 2px solid #222e27; border-top-color: #34c759; border-radius: 50%; display: none; animation: spin .7s linear infinite; }
  @keyframes spin { to { transform: rotate(360deg); } }

  .controls { padding: 8px 20px; background: #0e1412; border-bottom: 1px solid #222e27; display: flex; align-items: center; gap: 18px; flex-wrap: wrap; }
  .controls label { color: #4d6659; display: flex; align-items: center; gap: 6px; font-size: 12px; }
  .controls select { background: #151c18; color: #96b8a2; border: 1px solid #222e27; border-radius: 4px; padding: 3px 7px; font-size: 12px; }
  input[type=range] { width: 160px; accent-color: #34c759; cursor: pointer; }
  #convVal { color: #34c759; font-weight: 700; min-width: 26px; display: inline-block; }
  #rowCount { color: #e0ece4; font-weight: 600; }

  /* ── Summary bar ──────────────────────────────────────────── */
  .summary { display: flex; border-bottom: 1px solid #222e27; }
  .stat { padding: 10px 22px; border-right: 1px solid #222e27; }
  .stat-label { color: #4d6659; font-size: 10px; text-transform: uppercase; letter-spacing: .8px; margin-bottom: 1px; }
  .stat-value { font-size: 18px; font-weight: 700; color: #e0ece4; }
  .pos { color: #34c759; } .neg { color: #f0606e; }

  /* ── Sort bar ─────────────────────────────────────────────── */
  .sortbar { padding: 8px 20px; display: flex; gap: 6px; align-items: center; flex-wrap: wrap; border-bottom: 1px solid #222e27; background: #0e1412; }
  .sortbar button { background: #151c18; color: #4d6659; border: 1px solid #222e27; border-radius: 4px; padding: 4px 8px; font-size: 11px; font-family: inherit; cursor: pointer; transition: color .1s, border-color .1s; }
  .sortbar button:hover, .sortbar button.sorted { color: #34c759; border-color: #34c759; background: #182419; }
  .sortbar .sep { color: #222e27; font-size: 16px; }
  .odds-note { margin-left: auto; color: #4d6659; font-size: 11px; }

  /* ── Card list ────────────────────────────────────────────── */
  .cards { padding: 10px 20px 40px; display: flex; flex-direction: column; gap: 5px; }
  .empty { text-align: center; padding: 60px; color: #4d6659; }

  .play-card {
    display: grid;
    grid-template-columns: 68px 1fr auto;
    background: #111814;
    border: 1px solid #1e2921;
    border-radius: 8px;
    cursor: pointer;
    overflow: hidden;
    transition: border-color .12s, background .12s;
    min-height: 76px;
  }
  .play-card:hover { border-color: #2a4a34; background: #141f18; }

  /* Left conviction column */
  .conv-col {
    display: flex; align-items: center; justify-content: center;
    padding: 0 14px; border-right: 1px solid #1e2921; background: #0e1412; flex-shrink: 0;
  }
  .conv-num { font-size: 32px; font-weight: 800; line-height: 1; letter-spacing: -1px; }

  /* Middle content */
  .card-mid { padding: 12px 16px; display: flex; flex-direction: column; justify-content: center; gap: 4px; min-width: 0; }

  /* Market title — dominant element */
  .card-title { color: #d8ebe0; font-weight: 600; font-size: 14px; line-height: 1.3; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; margin-bottom: 1px; }

  /* Single plain-text meta line */
  .card-meta { font-size: 11px; color: #5a7a68; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
  .card-meta .hl { color: #8aa898; }
  .card-meta .up   { color: #34c759; }
  .card-meta .down { color: #f0606e; }
  .card-meta .gold { color: #e6a817; }

  /* Sub-row: outcome + stats as plain text */
  .card-sub { display: flex; align-items: center; gap: 0; font-size: 11px; color: #5a7a68; flex-wrap: nowrap; overflow: hidden; }
  .card-sub .sep { margin: 0 5px; color: #2a3a30; }
  .side-yes   { color: #34c759; font-weight: 700; }
  .side-no    { color: #f0606e; font-weight: 700; }
  .side-multi { color: #e6a817; font-weight: 700; }
  .sub-pos  { color: #34c759; }
  .sub-neg  { color: #f0606e; }

  /* Right position column */
  .card-right {
    display: flex; flex-direction: column; align-items: flex-end; justify-content: center;
    padding: 10px 16px; border-left: 1px solid #1e2921; gap: 4px; flex-shrink: 0;
  }
  .shares-tag { font-size: 10px; color: #4d6659; white-space: nowrap; }
  .value-tag  { font-size: 13px; font-weight: 700; color: #d8ebe0; white-space: nowrap; }
  .price-badge {
    background: #182d20; color: #34c759; border: 1px solid #234830;
    border-radius: 5px; padding: 3px 11px; font-weight: 700; font-size: 13px;
    white-space: nowrap; letter-spacing: .2px;
  }

  /* ── Shared small elements ───────────────────────────────── */
  .cat-badge { display: inline-block; background: #151c18; color: #5ab878; border: 1px solid #234830; border-radius: 3px; padding: 1px 6px; font-size: 10px; }
  .chip { background: #151c18; border: 1px solid #222e27; border-radius: 4px; padding: 2px 7px; font-size: 11px; white-space: nowrap; color: #96b8a2; }

  /* ── Drawer ───────────────────────────────────────────────── */
  .modal-backdrop { position: fixed; inset: 0; background: rgba(4,8,6,.85); display: none; z-index: 100; }
  .modal-backdrop.open { display: block; }
  .drawer { position: fixed; top: 0; right: 0; width: min(1020px, 96vw); height: 100vh; background: #111814; border-left: 1px solid #222e27; box-shadow: -20px 0 50px rgba(0,0,0,.5); overflow-y: auto; }
  .drawer-head { position: sticky; top: 0; background: #0e1412; border-bottom: 1px solid #222e27; padding: 14px 18px; z-index: 1; }
  .drawer-kicker { display: flex; gap: 8px; align-items: center; margin-bottom: 6px; }
  .drawer-title { color: #d8ebe0; font-size: 16px; font-weight: 700; margin-bottom: 4px; }
  .drawer-sub { color: #4d6659; font-size: 11px; }
  .drawer-close { position: absolute; top: 12px; right: 14px; background: #151c18; color: #8aa898; border: 1px solid #222e27; border-radius: 4px; padding: 4px 10px; cursor: pointer; font-size: 12px; }
  .drawer-close:hover { color: #34c759; border-color: #34c759; }

  /* Drawer stat line (single row under header) */
  .drawer-statline { padding: 10px 18px; border-bottom: 1px solid #222e27; font-size: 12px; color: #5a7a68; display: flex; gap: 0; flex-wrap: wrap; }
  .drawer-statline .dsl-sep { margin: 0 8px; color: #2a3a30; }
  .drawer-statline .dsl-hl  { color: #8aa898; }
  .drawer-statline .dsl-pos { color: #34c759; }
  .drawer-statline .dsl-neg { color: #f0606e; }

  /* Alignment summary row */
  .drawer-align { display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 8px; padding: 12px 18px; border-bottom: 1px solid #222e27; }
  .card-metric { min-width: 0; }
  .metric-label { color: #4d6659; font-size: 10px; text-transform: uppercase; letter-spacing: .6px; margin-bottom: 2px; }
  .metric-value { color: #d8ebe0; font-weight: 700; white-space: nowrap; }

  .drawer-body { padding: 0 18px 30px; }
  .wallet-section { margin-top: 16px; }
  .section-title { color: #d8ebe0; font-size: 13px; font-weight: 700; margin-bottom: 8px; }
  .modal-sort { display: flex; gap: 6px; flex-wrap: wrap; margin-bottom: 10px; }
  .modal-sort button { background: #151c18; color: #4d6659; border: 1px solid #222e27; border-radius: 4px; padding: 4px 8px; font: inherit; font-size: 11px; cursor: pointer; }
  .modal-sort button:hover { color: #34c759; border-color: #34c759; }

  .wallet-row { border: 1px solid #1e2921; border-radius: 6px; padding: 10px; margin-bottom: 7px; background: #0e1412; }
  .wallet-row.sharp { border-color: #234830; }
  .wallet-row-head { display: flex; justify-content: space-between; gap: 12px; flex-wrap: wrap; margin-bottom: 8px; }
  .wallet-address { color: #4d6659; font-size: 11px; overflow-wrap: anywhere; }
  .wallet-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(120px, 1fr)); gap: 7px 14px; }

  @media (max-width: 640px) {
    .play-card { grid-template-columns: 58px 1fr; }
    .card-right { display: none; }
    .drawer-panels { grid-template-columns: 1fr; }
  }
</style>
</head>
<body>

<header>
  <h1>Polymarket Tracker</h1>
  <span class="badge" id="walletCount">– wallets</span>
  <div id="spinner"></div>
  <div id="meta">–</div>
  <a href="/backtest" style="margin-left:auto;font-size:11px;color:#34c759;opacity:.6;border:1px solid #34c759;border-radius:8px;padding:2px 8px;white-space:nowrap">Backtest →</a>
  <a href="/export" style="font-size:11px;color:#e6a817;border:1px solid #e6a817;border-radius:8px;padding:2px 8px;white-space:nowrap;opacity:.75" title="Download Excel workbook">↓ Export XLSX</a>
</header>

<div class="controls">
  <label>Min conviction&nbsp;<span id="convVal">1</span>
    <input type="range" id="convSlider" min="1" max="100" value="1">
  </label>
  <label>Category <select id="categoryFilter"><option value="">All</option></select></label>
  <label>Wallet <select id="walletFilter"><option value="">All</option></select></label>
  <label>Resolves
    <select id="resolutionFilter">
      <option value="">All loaded</option>
      <option value="1">24h</option>
      <option value="2">48h</option>
    </select>
  </label>
  <span style="color:#3a6040">Showing <span id="rowCount">–</span> positions</span>
</div>

<div class="summary" id="summary"></div>

<div class="sortbar">
  <button onclick="sortBy('conviction')" id="th-conviction" class="sorted">Conviction</button>
  <button onclick="sortBy('tail_stake')" id="th-tail-stake">Tail</button>
  <button onclick="sortBy('sharp_wallet_count')" id="th-sharp-wallet-count">Aligned</button>
  <button onclick="sortBy('net_sharp_alignment')" id="th-net-sharp-alignment">Net</button>
  <span class="sep">|</span>
  <button onclick="sortBy('cur_price')" id="th-cur-price">Price</button>
  <button onclick="sortBy('size_usd')" id="th-size-usd">Invested</button>
  <button onclick="sortBy('pct_portfolio')" id="th-pct-portfolio">Port %</button>
  <span class="sep">|</span>
  <button onclick="sortBy('category_win_rate_pct')" id="th-category-win-rate-pct">Cat Win %</button>
  <button onclick="sortBy('category_roi_pct')" id="th-category-roi-pct">Cat ROI</button>
  <span class="odds-note">Odds -250 to +250</span>
</div>

<div class="cards" id="cards"></div>

<!-- Detail drawer -->
<div class="modal-backdrop" id="positionModal" onclick="backdropClose(event)">
  <div class="drawer">
    <div class="drawer-head">
      <button class="drawer-close" onclick="closePositionModal()">&#10005; Close</button>
      <div class="drawer-kicker" id="modalKicker"></div>
      <div class="drawer-title" id="modalTitle"></div>
      <div class="drawer-sub" id="modalSub"></div>
    </div>

    <!-- Key stat line -->
    <div class="drawer-statline" id="drawerStatline"></div>

    <!-- Alignment summary -->
    <div class="drawer-align" id="modalSummary"></div>

    <div class="drawer-body">
      <div class="modal-sort">
        <button onclick="sortModal('position_size')">Exposure</button>
        <button onclick="sortModal('total_pl')">P/L</button>
        <button onclick="sortModal('wallet_category_roi_pct')">Cat ROI</button>
        <button onclick="sortModal('wallet_historical_win_rate')">Win %</button>
        <button onclick="sortModal('portfolio_size')">Portfolio</button>
        <button onclick="sortModal('wallet_conviction_contribution')">Contribution</button>
      </div>
      <div class="wallet-section">
        <div class="section-title">Aligned wallets</div>
        <div id="alignedWallets"></div>
      </div>
      <div class="wallet-section">
        <div class="section-title">Opposing wallets</div>
        <div id="opposingWallets"></div>
      </div>
    </div>
  </div>
</div>

<script>
  let allData = [];
  let renderedData = [];
  let activeCard = null;
  let modalSortKey = 'position_size';
  let modalSortDir = -1;
  let sortKey = 'conviction';
  let sortDir = -1;
  const REFRESH = {{ refresh }} * 1000;

  const SPORT_ICONS = {
    'basketball': '&#127936;', 'nba': '&#127936;', 'wnba': '&#127936;',
    'football': '&#127944;', 'nfl': '&#127944;',
    'baseball': '&#9918;', 'mlb': '&#9918;',
    'hockey': '&#127944;', 'nhl': '&#127944;',
    'soccer': '&#9917;', 'fifa': '&#9917;',
    'tennis': '&#127934;',
    'mma': '&#129354;', 'ufc': '&#129354;', 'boxing': '&#129354;',
    'golf': '&#9971;',
    'esports': '&#127918;', 'gaming': '&#127918;', 'league': '&#127918;',
    'f1': '&#127950;', 'formula': '&#127950;',
    'cricket': '&#127951;',
    'rugby': '&#127944;',
  };

  function sportIcon(cat) {
    const lc = (cat || '').toLowerCase();
    for (const [k, v] of Object.entries(SPORT_ICONS)) {
      if (lc.includes(k)) return v;
    }
    return '&#127942;';
  }

  function convColor(c) {
    // hue sweeps 352° (red) → 142° (green) as conviction rises 0 → 100
    const hue = Math.round(-8 + 150 * (c / 100));
    return `hsl(${hue}, 68%, 48%)`;
  }

  function fmt(n, d=0) {
    return (n || 0).toLocaleString('en-US', {minimumFractionDigits: d, maximumFractionDigits: d});
  }

  function sign(n) { return (n || 0) >= 0 ? '+' : ''; }
  function pcls(n) { return (n || 0) >= 0 ? 'pos' : 'neg'; }

  function esc(v) {
    return String(v ?? '').replace(/[&<>"']/g, c =>
      ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c]));
  }

  function syncFilter(id, values) {
    const el = document.getElementById(id);
    const cur = el.value;
    el.innerHTML = '<option value="">All</option>' +
      values.map(v => `<option value="${esc(v)}">${esc(v)}</option>`).join('');
    if (values.includes(cur)) el.value = cur;
  }

  function alignedNames(row) {
    const ws = row.position_details?.aligned_wallets || [];
    if (ws.length) return [...new Set(ws.map(w => w.wallet).filter(Boolean))];
    return row.wallet ? [row.wallet] : [];
  }

  function walletMatches(row, w) { return !w || alignedNames(row).includes(w); }
  function walletOptions(rows) { return [...new Set(rows.flatMap(r => alignedNames(r)))].sort(); }

  function resolutionMatches(row, days) {
    if (!days) return true;
    const raw = row.resolution_time || row.end_date;
    if (!raw) return true;
    const end = new Date(raw);
    if (isNaN(end)) return true;
    const now = new Date();
    return end >= now && end <= new Date(+now + days * 86400000);
  }

  function backdropClose(e) { if (e.target.id === 'positionModal') closePositionModal(); }
  function closePositionModal() { document.getElementById('positionModal').classList.remove('open'); }

  function dm(label, value, cls='') {
    return `<div class="card-metric"><div class="metric-label">${label}</div><div class="metric-value ${cls}">${value}</div></div>`;
  }

  function dstat(label, val, sub='') {
    return `<div>
      <div class="d-stat-label">${label}</div>
      <div class="d-stat-val">${val}</div>
      ${sub ? `<div class="d-stat-sub">${sub}</div>` : ''}
    </div>`;
  }

  function sortWalletRows(rows) {
    return [...rows].sort((a, b) => {
      const av = a[modalSortKey], bv = b[modalSortKey];
      if (typeof av === 'string') return av.localeCompare(bv) * modalSortDir;
      return ((av||0) - (bv||0)) * modalSortDir;
    });
  }

  function renderWalletRows(rows) {
    if (!rows.length) return '<div style="color:#4d6659;padding:12px">None</div>';
    return sortWalletRows(rows).map(w => {
      const pc = (w.total_pl||0) >= 0 ? 'pos' : 'neg';
      const sc = w.side === 'Yes' ? 'side-yes' : 'side-no';
      return `<div class="wallet-row ${w.sharp_for_category ? 'sharp' : ''}">
        <div class="wallet-row-head">
          <div>
            <span class="chip">${esc(w.wallet)}</span>&nbsp;
            <span class="${sc}">${esc(w.side)}</span>
            ${w.sharp_for_category ? '&nbsp;<span class="cat-badge">sharp</span>' : ''}
            <div class="wallet-address">${esc(w.addr)}</div>
          </div>
          <div class="${pc}">${(w.total_pl||0)>=0?'+':'-'}$${fmt(Math.abs(w.total_pl||0))} &middot; ${sign(w.roi_pct)}${(w.roi_pct||0).toFixed(1)}%</div>
        </div>
        <div class="wallet-grid">
          ${dm('Entry',`${w.entry_price}&#162; <span style="color:#3a6040">${esc(w.entry_odds)}</span>`)}
          ${dm('Current',`${w.current_price}&#162; <span style="color:#3a6040">${esc(w.current_odds)}</span>`)}
          ${dm('Exposure',`$${fmt(w.position_size)}`)}
          ${dm('Shares',fmt(w.shares||0,2))}
          ${dm('Cost/Value',`$${fmt(w.cost_basis)} / $${fmt(w.current_value)}`)}
          ${dm('Unrealized',`${(w.unrealized_pl||0)>=0?'+':'-'}$${fmt(Math.abs(w.unrealized_pl||0))}`,pcls(w.unrealized_pl))}
          ${dm('Realized',`${(w.realized_pl||0)>=0?'+':'-'}$${fmt(Math.abs(w.realized_pl||0))}`,pcls(w.realized_pl))}
          ${dm('Portfolio',`$${fmt(w.wallet_total_portfolio_value||0)}`)}
          ${dm('% Port',`${(w.portfolio_pct||0).toFixed(2)}%`)}
          ${dm('Avg Position',`$${fmt(w.wallet_avg_position_size||0)}`)}
          ${dm('Size x',`${(w.position_size_multiple||0).toFixed(2)}x`)}
          ${dm('Cat Open',`$${fmt(w.wallet_category_portfolio_value||0)}`)}
          ${dm('Cat Volume',`$${fmt(w.wallet_category_total_volume||0)}`)}
          ${dm('Cat P/L',`${(w.wallet_category_pl||0)>=0?'+':'-'}$${fmt(Math.abs(w.wallet_category_pl||0))}`,pcls(w.wallet_category_pl))}
          ${dm('Cat ROI',`${sign(w.wallet_category_roi_pct)}${(w.wallet_category_roi_pct||0).toFixed(1)}%`)}
          ${dm('Hist Win',`${(w.wallet_historical_win_rate||0).toFixed(1)}%`)}
          ${dm('Source',esc(w.wallet_category_stats_source||'-'))}
          ${dm('Score',w.position_conviction||0)}
          ${dm('Contrib.',w.wallet_conviction_contribution||0)}
          ${w.source_link ? dm('Link',`<a href="${esc(w.source_link)}" target="_blank" onclick="event.stopPropagation()">Open &#8599;</a>`) : ''}
        </div>
      </div>`;
    }).join('');
  }

  function sl(text, cls='') {
    return `<span class="${cls}">${text}</span>`;
  }

  function renderModal() {
    if (!activeCard) return;
    const r = activeCard;
    const details = r.position_details || {};
    const s = details.summary || {};
    const aligned = details.aligned_wallets || [];
    const opposing = details.opposing_wallets || [];

    const isMulti = r.outcome === 'Multi';
    const outcome = isMulti ? 'Multi' : (s.selected_side || r.outcome || '');
    const sideCls = isMulti ? 'side-multi' : (outcome === 'Yes' ? 'side-yes' : 'side-no');
    const entryP  = r.avg_price || r.entry_price || 0;
    const curP    = r.cur_price || r.current_price || 0;
    const roi     = r.category_roi_pct || 0;
    const wr      = (r.category_win_rate_pct || 0).toFixed(1);
    const wl      = r.category_wins_losses || `${r.category_wins||'?'}-${r.category_losses||'?'}`;
    const mult    = (r.position_size_multiple || 1).toFixed(2);
    const roiCls  = roi >= 0 ? 'dsl-pos' : 'dsl-neg';

    document.getElementById('modalKicker').innerHTML =
      `<span class="cat-badge">${esc(s.category||r.category)}</span>
       <span class="${sideCls}" style="margin-left:2px">${esc(outcome)}</span>`;
    document.getElementById('modalTitle').textContent = s.market_title || r.market || '';
    if (isMulti && r.sub_markets && r.sub_markets.length > 1) {
      document.getElementById('modalSub').innerHTML =
        `Resolves ${s.resolution_time||r.end_date||'–'} &nbsp;&middot;&nbsp; `
        + r.sub_markets.map(m => `<span class="chip">${esc(m)}</span>`).join(' ');
    } else {
      document.getElementById('modalSub').textContent =
        `Resolves ${s.resolution_time||r.end_date||'–'}`;
    }

    const sep = `<span class="dsl-sep">&middot;</span>`;
    const priceStr = isMulti ? '' :
      sep + `${sl(fmt(entryP,0)+'&#162;')} &rarr; ${sl(fmt(curP,0)+'&#162;','dsl-hl')}`;
    document.getElementById('drawerStatline').innerHTML =
      `${sl('$'+fmt(r.size_usd||r.row_cost||0),'dsl-hl')} invested`
      + sep + `${sl(mult+'x','dsl-hl')} bet size`
      + priceStr
      + sep + `cat ROI ${sl(sign(roi)+roi.toFixed(1)+'%', roiCls)}`
      + sep + `win ${sl(wr+'%','dsl-hl')} (${wl})`
      + sep + `tail ${sl('$'+fmt(r.tail_stake||0),'dsl-hl')}`;

    const ac = (s.net_sharp_alignment||0) >= 0 ? 'pos' : 'neg';
    document.getElementById('modalSummary').innerHTML =
      dm('Aligned', s.aligned_sharp_wallet_count||0) +
      dm('Aligned exp.', `$${fmt(s.total_aligned_exposure||0)}`) +
      dm('Avg cat ROI', `${sign(s.average_aligned_category_roi)}${(s.average_aligned_category_roi||0).toFixed(1)}%`) +
      dm('Opposing', s.opposing_sharp_wallet_count||0) +
      dm('Opposing exp.', `$${fmt(s.opposing_exposure||0)}`) +
      dm('Net alignment', `${sign(s.net_sharp_alignment)}${s.net_sharp_alignment||0}`, ac) +
      dm('Net exposure', `${sign(s.net_sharp_exposure)}$${fmt(Math.abs(s.net_sharp_exposure||0))}`, ac) +
      dm('Score', s.final_conviction_score||r.conviction||0);

    document.getElementById('alignedWallets').innerHTML = renderWalletRows(aligned);
    document.getElementById('opposingWallets').innerHTML = renderWalletRows(opposing);
    document.getElementById('positionModal').classList.add('open');
  }

  function openCard(i) {
    activeCard = renderedData[i];
    modalSortKey = 'position_size'; modalSortDir = -1;
    renderModal();
  }

  function sortModal(key) {
    if (modalSortKey === key) modalSortDir *= -1; else { modalSortKey = key; modalSortDir = -1; }
    renderModal();
  }

  function render() {
    const minConv     = parseInt(document.getElementById('convSlider').value);
    const category    = document.getElementById('categoryFilter').value;
    const wallet      = document.getElementById('walletFilter').value;
    const resDays     = document.getElementById('resolutionFilter').value;
    document.getElementById('convVal').textContent = minConv;

    const filtered = allData.filter(r =>
      r.conviction >= minConv
      && (!category || r.category === category)
      && walletMatches(r, wallet)
      && resolutionMatches(r, resDays)
    );
    filtered.sort((a, b) => {
      const va = a[sortKey], vb = b[sortKey];
      if (typeof va === 'string') return va.localeCompare(vb) * sortDir;
      return ((va||0) - (vb||0)) * sortDir;
    });
    renderedData = filtered;
    document.getElementById('rowCount').textContent = filtered.length;

    const inv = filtered.reduce((s,r)=>s+(r.size_usd||0),0);
    const val = filtered.reduce((s,r)=>s+(r.current_value||0),0);
    const pnl = filtered.reduce((s,r)=>s+(r.row_pl||0),0);
    document.getElementById('summary').innerHTML = `
      <div class="stat"><div class="stat-label">Positions</div><div class="stat-value">${filtered.length}</div></div>
      <div class="stat"><div class="stat-label">Invested</div><div class="stat-value">$${fmt(inv)}</div></div>
      <div class="stat"><div class="stat-label">Current Value</div><div class="stat-value">$${fmt(val)}</div></div>
      <div class="stat"><div class="stat-label">Unrealised P/L</div><div class="stat-value ${pcls(pnl)}">${sign(pnl)}$${fmt(Math.abs(pnl))}</div></div>`;

    document.getElementById('walletCount').textContent =
      walletOptions(filtered).length + ' wallets';
    syncFilter('categoryFilter', [...new Set(allData.map(r=>r.category))].sort());
    syncFilter('walletFilter', walletOptions(allData));

    if (!filtered.length) {
      document.getElementById('cards').innerHTML =
        `<div class="empty">No positions above conviction ${minConv}</div>`;
      return;
    }

    document.getElementById('cards').innerHTML = filtered.map((r, i) => {
      const color  = convColor(r.conviction);
      const entryP = r.avg_price || r.entry_price || 0;
      const curP   = r.cur_price || r.current_price || 0;
      const curOdds = r.cur_odds || r.current_odds || '';

      const priceDelta = curP - entryP;
      const priceClass = priceDelta > 1 ? 'up' : priceDelta < -1 ? 'down' : '';

      const mult = r.position_size_multiple || 1;
      const multCls = mult >= 1.5 ? 'gold' : mult >= 1.0 ? '' : 'down';

      const roi = r.category_roi_pct || 0;
      const roiCls = roi >= 0 ? 'roi' : 'rneg';

      const wr  = (r.category_win_rate_pct || 0).toFixed(1);
      const wl  = r.category_wins_losses || `${r.category_wins||'?'}-${r.category_losses||'?'}`;
      const ico = sportIcon(r.category);

      const isMulti = r.outcome === 'Multi';
      const sideCls = isMulti ? 'side-multi' : (r.outcome === 'Yes' ? 'side-yes' : 'side-no');
      const names   = alignedNames(r);
      const walletLabel = names.length > 1 ? `${names.length} aligned` : esc(names[0] || r.wallet || '');

      const endRaw = r.resolution_time || r.end_date || '';
      let endLabel = endRaw ? endRaw.slice(0, 10) : '—';
      try {
        const d = new Date(endRaw);
        if (!isNaN(d)) endLabel = d.toLocaleDateString('en-US',{month:'short',day:'numeric'})
          + ' ' + d.toLocaleTimeString('en-US',{hour:'numeric',minute:'2-digit'});
      } catch(e){}

      const sep = `<span class="sep">&middot;</span>`;

      // Price moved colour
      const pDir = priceDelta > 1 ? 'up' : priceDelta < -1 ? 'down' : 'hl';
      // ROI colour
      const rCls = roi >= 0 ? 'up' : 'down';
      // Multiplier colour
      const mCls = mult >= 1.5 ? 'gold' : mult < 1.0 ? 'down' : 'hl';

      return `<article class="play-card" onclick="openCard(${i})">
        <div class="conv-col">
          <div class="conv-num" style="color:${color}">${r.conviction}</div>
        </div>
        <div class="card-mid">
          <div class="card-title" title="${esc(r.market)}">${esc(r.market)}</div>
          <div class="card-meta">
            ${ico} ${esc(r.category)}
            &nbsp;&middot;&nbsp;${esc(walletLabel)}
            &nbsp;&middot;&nbsp;${esc(endLabel)}
            &nbsp;&middot;&nbsp;<span class="hl">$${fmt(r.size_usd||r.row_cost||0)}</span>
            ${isMulti ? '' : `&nbsp;&middot;&nbsp;<span class="${pDir}">${fmt(entryP,0)}&#162;&rarr;${fmt(curP,0)}&#162;</span>`}
            &nbsp;&middot;&nbsp;<span class="${mCls}">${mult.toFixed(1)}x</span>
            &nbsp;&middot;&nbsp;ROI&nbsp;<span class="${rCls}">${sign(roi)}${roi.toFixed(1)}%</span>
          </div>
          <div class="card-sub">
            ${isMulti
              ? `<span class="side-multi">${r.event_market_count} markets</span>${sep}<span style="color:#4d6659;font-size:10px">${(r.sub_markets||[]).slice(0,2).map(m=>esc(m)).join(' &middot; ')+(r.sub_markets&&r.sub_markets.length>2?' &hellip;':'')}</span>`
              : `<span class="${sideCls}">${esc(r.outcome)}</span>${sep}<span>${fmt(curP,0)}&#162;&nbsp;<span style="color:#4d6659">${esc(curOdds)}</span></span>`
            }
            ${sep}<span>win ${wr}%&nbsp;<span style="color:#4d6659">(${wl})</span></span>
            ${sep}<span>${(r.pct_portfolio||0).toFixed(1)}% of portfolio</span>
            ${r.tail_stake ? `${sep}<span>tail $${fmt(r.tail_stake)}</span>` : ''}
          </div>
        </div>
        <div class="card-right">
          <div class="shares-tag">${isMulti ? (r.event_market_count+' markets') : fmt(r.shares||0,0)+' shares'}</div>
          <div class="value-tag">$${fmt(r.current_value||0)}</div>
          ${isMulti
            ? `<div class="price-badge" style="color:#e6a817;border-color:#4a3800;background:#1e1600">HEDGE</div>`
            : `<div class="price-badge">${fmt(curP,0)}&#162;</div>`
          }
        </div>
      </article>`;
    }).join('');
  }

  function sortBy(key) {
    document.querySelectorAll('.sortbar button').forEach(t => t.classList.remove('sorted'));
    if (sortKey === key) sortDir *= -1;
    else {
      sortKey = key;
      sortDir = ['conviction','tail_stake','sharp_wallet_count','net_sharp_alignment',
                 'category_roi_pct','category_win_rate_pct','size_usd','pct_portfolio',
                 'current_value','cur_price'].includes(key) ? -1 : 1;
    }
    const el = document.getElementById('th-' + key.replace(/_/g,'-'));
    if (el) el.classList.add('sorted');
    render();
  }

  async function fetchData() {
    document.getElementById('spinner').style.display = 'inline-block';
    try {
      const res  = await fetch('/api/positions');
      const json = await res.json();
      if (json.error) {
        document.getElementById('meta').textContent = 'Error: ' + json.error;
      } else {
        allData = json.data || [];
        document.getElementById('meta').textContent = 'Updated ' + (json.updated || '–');
      }
    } catch(e) {
      document.getElementById('meta').textContent = 'Fetch error — retrying';
    }
    document.getElementById('spinner').style.display = 'none';
    render();
  }

  document.getElementById('convSlider').addEventListener('input', render);
  document.getElementById('categoryFilter').addEventListener('change', render);
  document.getElementById('walletFilter').addEventListener('change', render);
  document.getElementById('resolutionFilter').addEventListener('change', render);
  fetchData();
  setInterval(fetchData, REFRESH);
</script>
</body>
</html>"""


# ── Entry point ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not WALLETS:
        raise SystemExit("No WALLETS configured in .env")

    # Seed the cache once synchronously so the first page load isn't empty
    print(f"Fetching initial positions for {len(WALLETS)} wallets…")
    try:
        initial = build_positions()
        with _lock:
            _cache["data"]    = initial
            _cache["updated"] = datetime.now(timezone.utc).strftime("%H:%M:%S UTC")
        print(f"Loaded {len(initial)} positions")
    except Exception as exc:
        print(f"Initial fetch error: {exc}")

    # Background refresh thread
    t = threading.Thread(target=refresh_loop, daemon=True)
    t.start()

    print(f"Dashboard running at http://localhost:{PORT}")
    app.run(host="0.0.0.0", port=PORT, debug=False)
